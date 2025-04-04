import io
import logging
import os
import sys

import openpyxl
from django.urls import reverse
from arabic_reshaper import arabic_reshaper
from bidi import get_display
from django.http import JsonResponse, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.views.decorators.http import require_GET
#from jdatetime import datetime as jdatetime
from jdatetime import datetime as jdatetime_datetime
from django.contrib import messages
from django.db.models import Q, Count
from django.db import models, transaction
from django import forms
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from .utils import is_admin_or_superuser
from .models import CarEntry, CarParts, EditLog, UserProfile, Document, Company, Complaint, CarCosts, Serialnumber, \
    FirstInspection
from .forms import CarEntryForm, CarPartsForm, CarCostsForm, CustomUserCreationForm, SearchLogForm, \
    FinalRegistrationForm, DocumentForm, ComplaintForm, SerialNumberform
from .forms import AdminChangePasswordForm, UserProfileEditForm
from django.contrib.auth import update_session_auth_hash
from .forms import AdminChangePasswordForm, UserProfileEditForm, AdminUserCreationForm, CustomPasswordChangeForm, FirstInspectionForm
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from django.db import IntegrityError
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
import re


def ensure_company_selected(request):
    sys.stdout.reconfigure(encoding='utf-8')
    print(f"ensure_company_selected: User={request.user}, Superuser={request.user.is_superuser}, Path={request.path}")
    if request.user.is_superuser:
        print(f"ensure_company_selected: Superuser {request.user.username} detected, no redirect")
        return None
    company = getattr(request, 'company', None)
    if not company and request.path not in ['/company_select/', '/login/', '/logout/', '/admin/']:
        print(f"ensure_company_selected: Redirecting to company_select for {request.user.username}, Company={company}")
        return redirect('company_select')
    return None
@login_required
def company_select(request):
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)

    if request.method == 'POST':
        company_id = request.POST.get('company_id')
        try:
            company = Company.objects.get(id=company_id, is_active=True)
            profile.company = company
            profile.save()
            messages.success(request, f"شرکت {company.name} با موفقیت انتخاب شد.")
            print(f"Company set for {request.user.username}: {company.name}")  # دیباگ
            return redirect('user_management')  # به جای home به user_management بره
        except Company.DoesNotExist:
            messages.error(request, "شرکت انتخاب‌شده معتبر نیست.")

    companies = Company.objects.filter(is_active=True)
    return render(request, 'core/company_select.html', {'companies': companies})


@login_required
def home(request):
    sys.stdout.reconfigure(encoding='utf-8')
    print(f"Home - User authenticated: {request.user.is_authenticated}")

    # چک کردن دسترسی و گرفتن شرکت
    if not request.user.is_superuser:
        if not hasattr(request.user, 'userprofile') or not request.user.userprofile.company:
            print("Home - User has no company, redirecting to company_select")
            return redirect('company_select')
        company = request.user.userprofile.company
        recent_cars = CarEntry.objects.filter(company=company).order_by('-accepted_at')[:5]
    else:
        company = None
        recent_cars = CarEntry.objects.order_by('-accepted_at')[:5]

    context = {
        'recent_cars': recent_cars
    }
    print(f"Home - Rendering with {len(recent_cars)} recent cars")
    return render(request, 'core/home.html', context)


class RoleChangeForm(forms.Form):
    role = forms.ChoiceField(
        choices=[('admin', 'ادمین'), ('regular', 'کاربر معمولی')],
        label="نقش",
        widget=forms.Select(attrs={'class': 'form-select'})
    )


def is_admin(user):
    try:
        return user.is_authenticated and user.userprofile.role == 'admin'
    except UserProfile.DoesNotExist:
        return False


@login_required
@user_passes_test(is_admin_or_superuser)
def user_management(request):
    redirect_response = ensure_company_selected(request)
    if redirect_response:
        return redirect_response
    company = request.company

    # اضافه کردن کاربر جدید
    if request.method == 'POST' and 'add_user' in request.POST:
        form = AdminUserCreationForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, "کاربر جدید با موفقیت اضافه شد.")
            return redirect('user_management')
    else:
        form = AdminUserCreationForm(request=request)

    # مدیریت اقدامات (حذف، فعال/غیرفعال)
    if request.method == 'POST' and 'action' in request.POST:
        user_id = request.POST.get('user_id')
        action = request.POST.get('action')
        user = get_object_or_404(User, id=user_id)

        if not request.user.is_superuser and user.userprofile.company != company:
            messages.error(request, "شما اجازه مدیریت این کاربر را ندارید.")
            return redirect('user_management')

        if action == 'delete':
            if user == request.user:
                messages.error(request, 'شما نمی‌توانید خودتان را حذف کنید.')
            else:
                user.delete()
                messages.success(request, 'کاربر با موفقیت حذف شد.')
        elif action == 'toggle_active':
            if user == request.user:
                messages.error(request, 'شما نمی‌توانید خودتان را غیرفعال کنید.')
            else:
                user.is_active = not user.is_active
                user.save()
                status = 'فعال' if user.is_active else 'غیرفعال'
                messages.success(request, f'کاربر با موفقیت {status} شد.')
        return redirect('user_management')

    if request.user.is_superuser:
        users = User.objects.all().order_by('-date_joined')
    else:
        users = User.objects.filter(userprofile__company=company).order_by('-date_joined')

    forms_list = []
    for user in users:
        try:
            role_form = RoleChangeForm(initial={'role': user.userprofile.role})
            forms_list.append((user, role_form))
        except UserProfile.DoesNotExist:
            forms_list.append((user, None))

    return render(request, 'core/user_management.html', {
        'users_with_forms': forms_list,
        'add_user_form': form,
    })


@login_required
@user_passes_test(is_admin_or_superuser)
def admin_change_password(request, user_id):
    redirect_response = ensure_company_selected(request)
    if redirect_response:
        return redirect_response
    company = request.company

    user = get_object_or_404(User, id=user_id)
    # چک کردن دسترسی
    if not request.user.is_superuser and user.userprofile.company != company:
        messages.error(request, "شما اجازه تغییر رمز این کاربر را ندارید.")
        return redirect('user_management')

    if request.method == 'POST':
        form = AdminChangePasswordForm(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            user.set_password(new_password)
            user.save()
            messages.success(request, f'رمز عبور کاربر {user.username} با موفقیت تغییر کرد.')
            return redirect('user_management')
        else:
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
    else:
        form = AdminChangePasswordForm()
    return render(request, 'core/admin_change_password.html', {
        'form': form,
        'user': user,
    })
@login_required
@user_passes_test(is_admin_or_superuser)
def edit_user_profile(request, user_id):
    redirect_response = ensure_company_selected(request)
    if redirect_response:
        return redirect_response
    company = request.company

    user = get_object_or_404(User, id=user_id)
    # چک کردن دسترسی
    if not request.user.is_superuser and user.userprofile.company != company:
        messages.error(request, "شما اجازه ویرایش این کاربر را ندارید.")
        return redirect('user_management')

    try:
        profile = user.userprofile
    except UserProfile.DoesNotExist:
        profile = UserProfile(user=user, company=company)  # شرکت پیش‌فرض

    if request.method == 'POST':
        form = UserProfileEditForm(request.POST, instance=profile, request=request)
        if form.is_valid():
            profile = form.save(commit=False)
            # چک کن که ادمین نقش خودش رو به regular تغییر نده
            if user == request.user and form.cleaned_data['role'] == 'regular' and profile.role != 'regular':
                messages.error(request, 'شما نمی‌توانید نقش خودتان را به کاربر معمولی تغییر دهید.')
            else:
                profile.save()
                messages.success(request, f'پروفایل کاربر {user.username} با موفقیت ویرایش شد.')
                return redirect('user_management')
        else:
            messages.error(request, 'اطلاعات واردشده نامعتبر است.')
    else:
        form = UserProfileEditForm(instance=profile, request=request)
        # اگه کاربر خودشو ویرایش می‌کنه، فیلد نقش رو غیرفعال کن
        if user == request.user:
            form.fields['role'].disabled = True
            form.fields['role'].help_text = 'شما نمی‌توانید نقش خودتان را تغییر دهید.'

    return render(request, 'core/edit_user_profile.html', {
        'form': form,
        'user': user,
    })
@login_required
@user_passes_test(is_admin_or_superuser)
def add_user(request):
    redirect_response = ensure_company_selected(request)
    if redirect_response:
        return redirect_response
    company = request.company

    if request.method == 'POST':
        form = AdminUserCreationForm(request.POST, request=request)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'کاربر {user.username} با موفقیت ثبت شد. رمز اولیه: {form.cleaned_data["national_code"]}')
            return redirect('user_management')
        else:
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
    else:
        form = AdminUserCreationForm(request=request)
    return render(request, 'core/add_user.html', {'form': form})

@login_required
def update_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            messages.success(request, 'رمز عبور شما با موفقیت تغییر کرد.')
            return redirect('home')  # یا هر صفحه‌ای که می‌خوای
        else:
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
    else:
        form = PasswordChangeForm(user=request.user)
    return render(request, 'core/update_password.html', {'form': form})

@login_required
def change_password(request):
    redirect_response = ensure_company_selected(request)
    if redirect_response and request.path != '/company_select/':
        return redirect_response

    profile, created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={'must_change_password': True, 'company': request.company}
    )
    if not profile.must_change_password:
        return redirect('home')

    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.POST)
        if form.is_valid():
            request.user.set_password(form.cleaned_data['new_password'])
            profile.must_change_password = False
            profile.save()
            request.user.save()
            update_session_auth_hash(request, request.user)
            messages.success(request, 'رمز عبور شما با موفقیت تغییر کرد.')
            return redirect('home')
        else:
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
    else:
        form = CustomPasswordChangeForm()
    return render(request, 'core/change_password.html', {'form': form})
@login_required
@user_passes_test(is_admin, login_url='home')
def edit_logs(request, acceptance_number='all'):
    form = SearchLogForm(request.GET or None)
    logs = EditLog.objects.all().order_by('-edited_at')
    car = None

    if not request.user.is_superuser:
        if not hasattr(request.user, 'userprofile') or not request.user.userprofile.company:
            messages.error(request, 'شما به هیچ شرکتی متصل نیستید.')
            return redirect('home')
        logs = logs.filter(company=request.user.userprofile.company)

    if form.is_valid() and form.cleaned_data['acceptance_number']:
        search_term = form.cleaned_data['acceptance_number']
        logs = logs.filter(car_entry__acceptance_number__icontains=search_term)

    if acceptance_number != 'all':
        car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)
        if not request.user.is_superuser and car.company != request.user.userprofile.company:
            messages.error(request, 'شما اجازه مشاهده لاگ‌های این خودرو را ندارید.')
            return redirect('home')
        logs = logs.filter(car_entry=car)

    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    try:
        page_logs = paginator.page(page_number)
    except PageNotAnInteger:
        page_logs = paginator.page(1)
    except EmptyPage:
        page_logs = paginator.page(paginator.num_pages)

    return render(request, 'core/edit_logs.html', {
        'car': car,
        'logs': page_logs,
        'form': form,
        'paginator': paginator,
    })
def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            profile, created = UserProfile.objects.get_or_create(user=user, defaults={'must_change_password': True})
            if profile.must_change_password:
                return redirect('change_password')
            return redirect('home')
    else:
        form = AuthenticationForm()
    return render(request, 'core/login.html', {'form': form})
def user_logout(request):
    logout(request)
    return redirect('login')

@login_required
def register_user(request):
    if not request.user.is_superuser:
        return redirect('registration_list')
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'کاربر جدید با موفقیت ثبت شد.')
            return redirect('registration_list')
    else:
        form = CustomUserCreationForm()
    return render(request, 'core/register_user.html', {'form': form})


@login_required
def select_car_for_parts(request):
    # برای سوپر یوزر همه خودروها، برای غیر سوپر یوزر فقط خودروهای شرکت خودش
    if request.user.is_superuser:
        cars_without_parts = CarEntry.objects.filter(parts__isnull=True)
    else:
        if not hasattr(request.user, 'userprofile') or not request.user.userprofile.company:
            messages.error(request, 'شما به هیچ شرکتی متصل نیستید.')
            return redirect('home')  # یا هر صفحه‌ای که مناسب باشه
        cars_without_parts = CarEntry.objects.filter(
            parts__isnull=True,
            company=request.user.userprofile.company
        )

    if request.method == 'POST':
        selected_car_number = request.POST.get('selected_car')
        if selected_car_number and cars_without_parts.filter(acceptance_number=selected_car_number).exists():

                # ریدایرکت با پارامتر next
            redirect_url = reverse('add_car_parts_step2', kwargs={'acceptance_number': selected_car_number})
            redirect_url += '?next=select_car_for_parts'
            return redirect(redirect_url)

        else:
            messages.error(request, 'لطفاً یک خودرو انتخاب کنید.')

    context = {
        'cars_without_parts': cars_without_parts,
    }
    return render(request, 'core/select_car_for_parts.html', context)


@login_required
def select_car_for_costs(request):
    if request.user.is_superuser:
        cars_without_costs = CarEntry.objects.filter(costs__isnull=True)
    else:
        if not hasattr(request.user, 'userprofile') or not request.user.userprofile.company:
            messages.error(request, 'شما به هیچ شرکتی متصل نیستید.')
            return redirect('home')
        cars_without_costs = CarEntry.objects.filter(
            costs__isnull=True,
            company=request.user.userprofile.company
        )

    if request.method == 'POST':
        selected_car_number = request.POST.get('selected_car')
        if selected_car_number and cars_without_costs.filter(acceptance_number=selected_car_number).exists():
            # ریدایرکت با پارامتر next
            redirect_url = reverse('add_car_costs_step3', kwargs={'acceptance_number': selected_car_number})
            redirect_url += '?next=select_car_for_costs'
            return redirect(redirect_url)
        else:
            messages.error(request, 'لطفاً یک خودرو معتبر انتخاب کنید.')

    context = {
        'cars_without_costs': cars_without_costs,
    }
    return render(request, 'core/select_car_for_costs.html', context)

# @login_required
# def add_document(request):
#     sys.stdout.reconfigure(encoding='utf-8')
#     selected_car = request.session.get('selected_car')
#     car = None
#     if selected_car:
#         car = get_object_or_404(CarEntry, acceptance_number=selected_car)
#
#     if request.method == 'POST':
#         form = DocumentForm(request.POST, request.FILES)
#         if form.is_valid():
#             document = form.save(commit=False)
#             document.created_by = request.user
#             if car:
#                 # پر کردن شماره موتور و شاسی از ماشین انتخاب‌شده
#                 document.engine_number = car.engine_number
#                 document.chassis_number = car.chassis_number  # اضافه کردن شماره شاسی
#                 document.car = car  # اتصال سند به خودرو
#             document.save()
#
#             # برای ثبت معکوس توی CarEntry
#             if car:
#                 car.document = document
#                 car.save(update_fields=['document'])  # آپدیت فقط document_id
#
#             messages.success(request, f"سند با شماره موتور {document.engine_number} با موفقیت ثبت شد.")
#             if 'selected_car' in request.session:
#                 del request.session['selected_car']
#             return redirect('registration_list')
#         else:
#             messages.error(request, 'لطفاً خطاها را برطرف کنید.')
#             print("Form errors:", form.errors)
#     else:
#         initial_data = {}
#         if car:
#             initial_data['engine_number'] = car.engine_number
#             initial_data['chassis_number'] = car.chassis_number  # مقدار اولیه برای شاسی
#         form = DocumentForm(initial=initial_data)
#     return render(request, 'core/add_document.html', {'form': form, 'car': car})#

@login_required
def add_complaint(request, acceptance_number):
    car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)




    # گرفتن مقصد بازگشت از پارامتر next یا سشن (اولویت با GET)
    next_url = request.GET.get('next', request.session.get('return_to', 'registration_list'))
    logger.debug(f"Next URL: {next_url}")

    # گرفتن پارامترهای سرچ از سشن
    search_params = request.session.get('search_params', {})
    logger.debug(f"Search params from session: {search_params}")

    if request.method == 'POST':
        form = ComplaintForm(request.POST, request.FILES)
        if form.is_valid():
            complaint = form.save(commit=False)
            complaint.car = car
         #   complaint.created_by = request.user
            if hasattr(request.user,
                       'userprofile') and request.user.userprofile.first_name and request.user.userprofile.last_name:
                complaint.created_by = f"{request.user.userprofile.first_name} {request.user.userprofile.last_name}"
            else:
                complaint.created_by = request.user.username

            complaint.company = car.company if hasattr(car, 'company') else None
            if not complaint.company:
                logger.error("No valid company for car or user")
                messages.error(request, "خودرو یا کاربر شرکت معتبری ندارد. لطفاً شرکت را تنظیم کنید.")
                return render(request, 'core/add_complaint.html',
                              {'form': form, 'car': car, 'next_url': next_url})

            complaint.save()
            messages.success(request, f"شکایت '{complaint.title}' برای خودرو {car.acceptance_number} ثبت شد.")
            # ساخت URL با توجه به next_url
            redirect_url = reverse(next_url)
            if next_url == 'registration_list' and search_params:
                query_string = '&'.join([f"{key}={value}" for key, value in search_params.items()])
                redirect_url += f"?{query_string}"
            logger.debug(f"Redirecting to: {redirect_url}")
            return HttpResponseRedirect(redirect_url)



        else:
            messages.error(request, "لطفاً خطاها را برطرف کنید.")
            return render(request, 'core/add_complaint.html', {'form': form, 'car': car, 'next_url': next_url})
    else:
        form = ComplaintForm()

    return render(request, 'core/add_complaint.html', {
        'form': form,
        'car': car,
        'next_url': 'registration_list',  # مقصد بعدی
        'search_params': search_params  # پارامترهای جستجو برای دکمه برگشت
    })


@login_required
def manage_complaints(request, acceptance_number):
    car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)

    # گرفتن شرکت کاربر و فیلتر کردن شکایت‌ها
    try:
        company = request.user.userprofile.company
        if not company:
            messages.error(request, "شما به شرکتی متصل نیستید.")
            return redirect('company_select')
        complaints = Complaint.objects.filter(car=car, company=company).order_by('-created_at')
    except AttributeError:
        messages.error(request, "پروفایل شما کامل نیست.")
        return redirect('company_select')

    # گرفتن مقصد بازگشت از پارامتر next یا سشن (اولویت با GET)
    next_url = request.GET.get('next', request.session.get('return_to', 'registration_list'))
    logger.debug(f"Next URL: {next_url}")

    # گرفتن پارامترهای جستجو از سشن
    search_params = request.session.get('search_params', {})
    logger.debug(f"Search params from session: {search_params}")

    if request.method == 'POST':
        # بررسی اینکه آیا درخواست برای ثبت شکایت جدید است یا مدیریت شکایت موجود
        if 'complaint_id' in request.POST:  # مدیریت شکایت موجود
            complaint_id = request.POST.get('complaint_id')
            action = request.POST.get('action')
            complaint = get_object_or_404(Complaint, id=complaint_id, car=car, company=company)

            if action == 'resolve':
                complaint.is_resolved = True
                complaint.save()
                messages.success(request, f"شکایت '{complaint.title}' رفع شد.")
            return redirect('manage_complaints', acceptance_number=car.acceptance_number)

        else:  # ثبت شکایت جدید
            form = ComplaintForm(request.POST, request.FILES)
            if form.is_valid():
                complaint = form.save(commit=False)
                complaint.car = car
                if hasattr(request.user, 'userprofile') and request.user.userprofile.first_name and request.user.userprofile.last_name:
                    complaint.created_by = f"{request.user.userprofile.first_name} {request.user.userprofile.last_name}"
                else:
                    complaint.created_by = request.user.username

                complaint.company = car.company if hasattr(car, 'company') else company
                if not complaint.company:
                    logger.error("No valid company for car or user")
                    messages.error(request, "خودرو یا کاربر شرکت معتبری ندارد. لطفاً شرکت را تنظیم کنید.")
                    return render(request, 'core/manage_complaints.html', {
                        'form': form,
                        'car': car,
                        'complaints': complaints,
                        'next_url': next_url,
                        'search_params': search_params
                    })

                complaint.save()
                messages.success(request, f"شکایت '{complaint.title}' برای خودرو {car.acceptance_number} ثبت شد.")
                # به‌جای هدایت، فرم را خالی کنیم و لیست شکایت‌ها را به‌روز کنیم
                form = ComplaintForm()  # فرم جدید خالی
                complaints = Complaint.objects.filter(car=car, company=company).order_by('-created_at')  # به‌روزرسانی لیست
                return render(request, 'core/manage_complaints.html', {
                    'form': form,
                    'car': car,
                    'complaints': complaints,
                    'next_url': next_url,
                    'search_params': search_params
                })
            else:
                messages.error(request, "لطفاً خطاها را برطرف کنید.")
                return render(request, 'core/manage_complaints.html', {
                    'form': form,
                    'car': car,
                    'complaints': complaints,
                    'next_url': next_url,
                    'search_params': search_params
                })

    # درخواست GET: نمایش فرم و لیست شکایت‌ها
    form = ComplaintForm()

    return render(request, 'core/manage_complaints.html', {
        'form': form,
        'car': car,
        'complaints': complaints,
        'next_url': next_url,
        'search_params': search_params
    })
@login_required
def document_list(request):
    sys.stdout.reconfigure(encoding='utf-8')

    # بررسی شرکت کاربر
    if not request.user.is_superuser:
        try:
            if not hasattr(request.user, 'userprofile') or not request.user.userprofile.company:
                logger.warning("User has no profile or company, redirecting to company_select")
                messages.error(request, 'شما به هیچ شرکتی متصل نیستید.')
                return redirect('company_select')
            company = request.user.userprofile.company
            documents = Document.objects.filter(company=company).order_by('-delivery_date_gregorian')
            logger.debug(f"Company: {company}, Initial queryset count: {documents.count()}")
        except Exception as e:
            logger.error(f"Error in company check: {str(e)}")
            messages.error(request, f'خطا در دسترسی به شرکت: {str(e)}')
            return redirect('company_select')
    else:
        company = None
        try:
            documents = Document.objects.all().order_by('-delivery_date_gregorian')
            logger.debug(f"Superuser, Initial queryset count: {documents.count()}")
        except Exception as e:
            logger.error(f"Error fetching documents for superuser: {str(e)}")
            messages.error(request, f'خطا در بارگذاری داده‌ها: {str(e)}')
            documents = Document.objects.none()

    # گرفتن پارامترهای جستجو
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    engine_number = request.GET.get('engine_number', '')
    chassis_number = request.GET.get('chassis_number', '')
    car_status = request.GET.get('car_status', '')
    acceptance_number = request.GET.get('acceptance_number', '')
    page_number = request.GET.get('page', '')

    # اعتبارسنجی تاریخ‌ها
    date_pattern = r'^\d{4}/\d{2}/\d{2}$'
    if date_from and not re.match(date_pattern, date_from):
        messages.error(request, 'فرمت "از تاریخ" اشتباه است. لطفاً از فرمت 1403/12/15 استفاده کنید.')
        date_from = ''
    if date_to and not re.match(date_pattern, date_to):
        messages.error(request, 'فرمت "تا تاریخ" اشتباه است. لطفاً از فرمت 1403/12/15 استفاده کنید.')
        date_to = ''
    if date_from and date_to:
        try:
            start_jdate = jdatetime_datetime.strptime(date_from, '%Y/%m/%d')
            end_jdate = jdatetime_datetime.strptime(date_to, '%Y/%m/%d')
            if start_jdate > end_jdate:
                messages.error(request, '"از تاریخ" نمی‌تواند بزرگ‌تر از "تا تاریخ" باشد.')
                date_from = ''
                date_to = ''
        except ValueError:
            messages.error(request, 'تاریخ‌ها نامعتبر هستند.')
            date_from = ''
            date_to = ''

    # اعمال فیلترها
    try:
        if engine_number and len(engine_number) <= 50:
            documents = documents.filter(engine_number__icontains=engine_number)
        elif engine_number:
            messages.error(request, 'شماره موتور نمی‌تواند بیشتر از 50 کاراکتر باشد.')

        if chassis_number and len(chassis_number) <= 50:
            documents = documents.filter(chassis_number__icontains=chassis_number)
        elif chassis_number:
            messages.error(request, 'شماره شاسی نمی‌تواند بیشتر از 50 کاراکتر باشد.')

        if date_from:
            try:
                start_date_gregorian = jdatetime_datetime.strptime(date_from, '%Y/%m/%d').togregorian()
                documents = documents.filter(delivery_date_gregorian__gte=start_date_gregorian)
            except ValueError:
                messages.error(request, 'تاریخ شروع نامعتبر است.')

        if date_to:
            try:
                end_date_gregorian = jdatetime_datetime.strptime(date_to, '%Y/%m/%d').togregorian()
                documents = documents.filter(delivery_date_gregorian__lte=end_date_gregorian)
            except ValueError:
                messages.error(request, 'تاریخ پایان نامعتبر است.')

        if car_status:
            if car_status == 'connected':
                documents = documents.filter(car__isnull=False)
            elif car_status == 'not_connected':
                documents = documents.filter(car__isnull=True)

        if acceptance_number:
            documents = documents.filter(car__acceptance_number__icontains=acceptance_number)
    except Exception as e:
        logger.error(f"Error applying filters: {str(e)}")
        messages.error(request, f'خطا در اعمال فیلترها: {str(e)}')

    # ذخیره پارامترهای جستجو در سشن
    search_params = {
        'date_from': date_from,
        'date_to': date_to,
        'engine_number': engine_number,
        'chassis_number': chassis_number,
        'car_status': car_status,
        'acceptance_number': acceptance_number,
        'page': page_number
    }
    try:
        request.session['search_params'] = {k: v for k, v in search_params.items() if v}
        logger.debug("Search params saved to session")
    except Exception as e:
        logger.error(f"Error saving session: {str(e)}")
        messages.error(request, f'خطا در ذخیره سشن: {str(e)}')

    # صفحه‌بندی
    paginator = Paginator(documents, 10)
    try:
        page_obj = paginator.page(page_number if page_number else 1)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    except Exception as e:
        logger.error(f"Pagination error: {str(e)}")
        messages.error(request, f'خطا در صفحه‌بندی: {str(e)}')
        page_obj = paginator.page(1)

    # آماده‌سازی برای رندر
    context = {
        'page_obj': page_obj,
        'date_from': date_from,
        'date_to': date_to,
        'engine_number': engine_number,
        'chassis_number': chassis_number,
        'car_status': car_status,
        'acceptance_number': acceptance_number,
    }
    return render(request, 'core/document_list.html', context)
@login_required
def add_document(request):
    sys.stdout.reconfigure(encoding='utf-8')

    # گرفتن car_id از GET یا session
    car_id = request.GET.get('car_id', None)
    selected_car_session = request.session.get('selected_car', None)
    initial_data = {}
    selected_car = None

    # اولویت با GET، بعد session
    if car_id:
        try:
            selected_car = CarEntry.objects.get(id=car_id)
            initial_data = {
                'engine_number': selected_car.engine_number,
                'chassis_number': selected_car.chassis_number,
            }
        except CarEntry.DoesNotExist:
            messages.error(request, 'خودرو انتخاب‌شده پیدا نشد.')
    elif selected_car_session:
        try:
            selected_car = CarEntry.objects.get(acceptance_number=selected_car_session)
            initial_data = {
                'engine_number': selected_car.engine_number,
                'chassis_number': selected_car.chassis_number,
            }
            # پاک کردن session بعد از استفاده
            if 'selected_car' in request.session:
                del request.session['selected_car']
        except CarEntry.DoesNotExist:
            messages.error(request, 'خودرو انتخاب‌شده از لیست پیدا نشد.')

    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES, request=request, initial=initial_data)
        if form.is_valid():
            document = form.save()
            messages.success(request, f"سند {document.engine_number} - {document.chassis_number} با موفقیت ثبت شد.")
            next_page = request.GET.get('next', 'document_list')
            if next_page == 'registration_list':
                search_params = request.session.get('search_params', {})
                query_string = '&'.join([f"{k}={v}" for k, v in search_params.items() if v])
                redirect_url = reverse('registration_list')
                if query_string:
                    redirect_url += f"?{query_string}"
                return redirect(redirect_url)
            return redirect(next_page)
        else:
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
            print("خطاهای فرم:", form.errors)
    else:
        form = DocumentForm(request=request, initial=initial_data)

    previous_page = request.GET.get('next', 'registration_list')
    previous_url = reverse(previous_page)
    if previous_page == 'registration_list':
        search_params = request.session.get('search_params', {})
        query_string = '&'.join([f"{k}={v}" for k, v in search_params.items() if v])
        if query_string:
            previous_url += f"?{query_string}"

    return render(request, 'core/add_document.html', {
        'form': form,
        'previous_url': previous_url,
        'selected_car': selected_car,
    })

@login_required
def edit_document(request, pk):
    sys.stdout.reconfigure(encoding='utf-8')
    document = get_object_or_404(Document, pk=pk)

    # پیش‌فرض: خودرو متصل به سند
    selected_car = document.car

    # گرفتن car_id از GET یا session (اولویت با این‌ها)
    car_id = request.GET.get('car_id', None)
    selected_car_session = request.session.get('selected_car', None)

    # بررسی دسترسی
    if not request.user.is_superuser:
        try:
            if document.company != request.user.userprofile.company:
                messages.error(request, "شما اجازه ویرایش این سند را ندارید.")
                return redirect('document_list')
        except AttributeError:
            messages.error(request, "پروفایل شما کامل نیست.")
            return redirect('company_select')

    # اگه car_id یا session باشه، اولویت می‌دیم
    if car_id:
        try:
            selected_car = CarEntry.objects.get(id=car_id)
        except CarEntry.DoesNotExist:
            messages.error(request, 'خودرو انتخاب‌شده پیدا نشد.')
    elif selected_car_session:
        try:
            selected_car = CarEntry.objects.get(acceptance_number=selected_car_session)
            if 'selected_car' in request.session:
                del request.session['selected_car']
        except CarEntry.DoesNotExist:
            messages.error(request, 'خودرو انتخاب‌شده از لیست پیدا نشد.')
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES, instance=document, request=request)
        if form.is_valid():
            # ثبت تغییرات
            changes = form.changed_data  # لیست فیلدهایی که تغییر کردن
            document = form.save()

            # ثبت لاگ ویرایش
            EditLog.objects.create(
                edit_type='DOCUMENT',  # نوع ویرایش رو به سند تغییر دادم
                car_entry=document.car,
                edited_by=request.user,
                company=document.company,
                changes=changes if changes else None
            )

            # پیام موفقیت
            messages.success(request, f"سند با شماره موتور {document.engine_number} با موفقیت ویرایش شد.")

            # بررسی صفحه بعدی
            next_page = request.GET.get('next', 'document_list')  # پیش‌فرض document_list
            if next_page == 'document_list':
                search_params = request.session.get('search_params', {})
                query_string = '&'.join([f"{k}={v}" for k, v in search_params.items() if v])
                redirect_url = reverse('document_list')
                if query_string:
                    redirect_url += f"?{query_string}"
                return redirect(redirect_url)
            if next_page == 'registration_list':
                search_params = request.session.get('search_params', {})
                query_string = '&'.join([f"{k}={v}" for k, v in search_params.items() if v])
                redirect_url = reverse('registration_list')
                if query_string:
                    redirect_url += f"?{query_string}"
                return redirect(redirect_url)
            return redirect(next_page)
        else:
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
            logger.error(f"Form errors: {form.errors}")
    else:
        form = DocumentForm(instance=document, request=request)

    # تنظیم صفحه قبلی برای بازگشت
    previous_page = request.GET.get('next', 'document_list')  # پیش‌فرض document_list
    previous_url = reverse(previous_page)
    if previous_page == 'document_list':
        search_params = request.session.get('search_params', {})
        query_string = '&'.join([f"{k}={v}" for k, v in search_params.items() if v])
        if query_string:
            previous_url += f"?{query_string}"
    if previous_page == 'registration_list':
        search_params = request.session.get('search_params', {})
        query_string = '&'.join([f"{k}={v}" for k, v in search_params.items() if v])
        if query_string:
            previous_url += f"?{query_string}"

    return render(request, 'core/edit_document.html', {
        'form': form,
        'document': document,
        'previous_url': previous_url,
        'selected_car': selected_car,
    })


@login_required
def update_serial_number(request, acceptance_number, car_id=None):
    sys.stdout.reconfigure(encoding='utf-8')
    car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)
    mode = request.GET.get('mode', 'add')
    serial_instance = Serialnumber.objects.filter(car=car).first()

    if mode not in ('add', 'edit'):
        mode = 'add'
    if mode == 'edit' and not serial_instance:
        messages.error(request, 'شماره سریالی برای ویرایش وجود ندارد.')
        mode = 'add'

    if request.method == 'POST':
        print("POST data:", request.POST)  # لاگ داده‌های متنی
        print("FILES data:", request.FILES)  # لاگ فایل‌ها
        form = SerialNumberform(request.POST, request.FILES, instance=serial_instance if mode == 'edit' else None)
        if form.is_valid():
            car_serialnumber = form.save(commit=False)
            car_serialnumber.car = car
            if hasattr(request.user,
                       'userprofile') and request.user.userprofile.first_name and request.user.userprofile.last_name:
                car_serialnumber.recorded_by = f"{request.user.userprofile.first_name} {request.user.userprofile.last_name}"
            else:
                car_serialnumber.recorded_by = request.user.username
            car_serialnumber.company = car.company
            if not car_serialnumber.company:
                messages.error(request, "این خودرو به هیچ شرکتی متصل نیست.")
                return redirect('update_serial_number', acceptance_number=acceptance_number)
            try:
                car_serialnumber.save()
                print("Serialnumber saved:", car_serialnumber.serial_number,
                      car_serialnumber.Registration_nnhk)  # لاگ بعد از ذخیره
                action = "ویرایش" if mode == 'edit' else "ثبت"
                messages.success(request, f"شماره سریال با شماره پذیرش {car.acceptance_number} با موفقیت {action} شد.")

                next_page = request.GET.get('next', 'registration_list')
                if next_page == 'registration_list':
                    search_params = request.session.get('search_params', {})
                    query_string = '&'.join([f"{k}={v}" for k, v in search_params.items() if v])
                    redirect_url = reverse('registration_list')
                    if query_string:
                        redirect_url += f"?{query_string}"
                    return redirect(redirect_url)
                return redirect(next_page)
            except IntegrityError:
                messages.error(request, 'خطا: شماره سریال این خودرو قبلاً ثبت شده است.')
            except Exception as e:
                messages.error(request, f"خطا در {action} شماره سریال: {str(e)}")
                return redirect('update_serial_number', acceptance_number=acceptance_number)
        else:
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
            print("Form errors:", form.errors)  # لاگ خطاها
    else:
        initial_data = {'serial_number': serial_instance.serial_number} if mode == 'edit' and serial_instance else {}
        form = SerialNumberform(instance=serial_instance if mode == 'edit' else None, initial=initial_data)

    previous_page = request.GET.get('next', 'registration_list')
    previous_url = reverse(previous_page)
    if previous_page == 'registration_list':
        search_params = request.session.get('search_params', {})
        query_string = '&'.join([f"{k}={v}" for k, v in search_params.items() if v])
        if query_string:
            previous_url += f"?{query_string}"

    return render(request, 'core/update_serial_number.html', {
        'form': form,
        'car': car,
        'previous_url': previous_url,
        'mode': mode,
    })
@login_required
def add_car_entry_step1(request):
    sys.stdout.reconfigure(encoding='utf-8')
    if request.method == 'POST':
        form = CarEntryForm(request.POST, request.FILES, request=request)
        if form.is_valid():
            car_entry = form.save(commit=False)
            # car_entry.accepted_by رو اینجا ست نمی‌کنیم چون توی save مدیریت می‌شه
            company = form.cleaned_data['company'] if request.user.is_superuser else request.user.userprofile.company
            car_entry.save(user=request.user, company=company)
            messages.success(request, f"خودرو با شماره پذیرش {car_entry.acceptance_number} توسط {car_entry.accepted_by} در شرکت {car_entry.company.name if car_entry.company else 'نامشخص'} با موفقیت ثبت شد")
            return redirect('home')
        else:
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
            print("Form errors:", form.errors)
    else:
        form = CarEntryForm(request=request)
    return render(request, 'core/add_car_entry_step1.html', {'form': form})




@login_required
def edit_car_entry(request, acceptance_number):
    sys.stdout.reconfigure(encoding='utf-8')
    car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)

    # چک کردن دسترسی شرکت
    if not request.user.is_superuser and car.company != request.user.userprofile.company:
        messages.error(request, 'شما اجازه ویرایش این خودرو را ندارید.')
        return redirect('registration_list')

    if request.method == 'POST':
        form = CarEntryForm(request.POST, request.FILES, instance=car, request=request)
        if form.is_valid():
            car_instance = form.save(commit=False)
            # پیدا کردن تغییرات
            changes = {}
            old_instance = CarEntry.objects.get(pk=car.pk)  # نسخه قدیمی
            for field in form.changed_data:
                old_value = getattr(old_instance, field)
                new_value = form.cleaned_data[field]
                if old_value != new_value:
                    changes[field] = {'old': str(old_value), 'new': str(new_value)}

            # تنظیم شرکت
            company = form.cleaned_data['company'] if request.user.is_superuser else request.user.userprofile.company
            if not company:
                messages.error(request, 'شرکت برای این خودرو یا کاربر مشخص نشده است.')
                return redirect('registration_list')

            car_instance.save(user=request.user, company=company)

            # ثبت لاگ با تغییرات
            if changes:
                EditLog.objects.create(
                    edit_type='CAR_ENTRY',
                    car_entry=car_instance,
                    edited_by=request.user,
                    company=company,
                    changes=changes
                )
            messages.success(request, f"اطلاعات خودرو با موفقیت توسط {request.user.username} ویرایش شد.")

            # بازیابی پارامترهای جستجو از سشن
            search_params = request.session.get('search_params', {})
            query_string = '&'.join([f"{key}={value}" for key, value in search_params.items() if value]) if search_params else ''
            redirect_url = reverse('registration_list')
            if query_string:
                redirect_url += f"?{query_string}"
            return redirect(redirect_url)
        else:
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
            print("Form errors:", form.errors)
    else:
        initial_data = {}
        if car.license_plate:
            try:
                parts = car.license_plate.split(' - ')
                if len(parts) == 4 and parts[0].startswith('ایران '):
                    initial_data['digits3'] = parts[0].replace('ایران ', '')
                    initial_data['digits2'] = parts[1]
                    initial_data['persian_letter'] = parts[2]
                    initial_data['digits1'] = parts[3]
            except Exception as e:
                print(f"Error parsing license_plate: {e}")
        if car.driver_license_plate:
            try:
                parts = car.driver_license_plate.split(' - ')
                if len(parts) == 4 and parts[0].startswith('ایران '):
                    initial_data['driver_digits3'] = parts[0].replace('ایران ', '')
                    initial_data['driver_digits2'] = parts[1]
                    initial_data['driver_persian_letter'] = parts[2]
                    initial_data['driver_digits1'] = parts[3]
            except Exception as e:
                print(f"Error parsing driver_license_plate: {e}")
        form = CarEntryForm(instance=car, initial=initial_data, request=request)

    return render(request, 'core/edit_car_entry.html', {'form': form, 'car': car})


@login_required
def add_car_parts_step2(request, acceptance_number):
    sys.stdout.reconfigure(encoding='utf-8')
    car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)

    if hasattr(car, 'parts') and car.parts:
        messages.error(request, 'قطعات این خودرو قبلاً ثبت شده است. لطفاً از ویرایش قطعات استفاده کنید.')
        return redirect('edit_car_parts', acceptance_number=car.acceptance_number)

    if request.method == 'POST':
        form = CarPartsForm(request.POST, request=request)
        if form.is_valid():
            car_parts = form.save(commit=False)
            car_parts.car = car
            # تنظیم recorded_by
            if hasattr(request.user,
                       'userprofile') and request.user.userprofile.first_name and request.user.userprofile.last_name:
                car_parts.recorded_by = f"{request.user.userprofile.first_name} {request.user.userprofile.last_name}"
            else:
                car_parts.recorded_by = request.user.username
            # تنظیم company
            car_parts.company = car.company
            if not car_parts.company:
                messages.error(request, "این خودرو به هیچ شرکتی متصل نیست.")
                return redirect('add_car_parts_step2', acceptance_number=acceptance_number)
            try:
                print("قبل از ذخیره:", car_parts.__dict__)
                car_parts.save()
                saved_parts = CarParts.objects.get(car=car)  # چک کن ذخیره شده
                print("بعد از ذخیره:", saved_parts.__dict__)
                messages.success(request,
                                 f"قطعات خودرو با شماره پذیرش {car.acceptance_number} توسط {car_parts.recorded_by} با موفقیت ثبت شد.")

                next_page = request.GET.get('next', 'registration_list')
                if next_page == 'registration_list':
                    search_params = request.session.get('search_params', {})
                    query_string = '&'.join([f"{k}={v}" for k, v in search_params.items() if v])
                    redirect_url = reverse('registration_list')
                    if query_string:
                        redirect_url += f"?{query_string}"
                    return redirect(redirect_url)
                return redirect(next_page)
            except IntegrityError:
                messages.error(request, 'خطا: قطعات برای این خودرو قبلاً ثبت شده است.')
            except Exception as e:
                print(f"خطا توی ذخیره‌سازی: {str(e)}")
                messages.error(request, f"خطا توی ثبت قطعات: {str(e)}")
                return redirect('add_car_parts_step2', acceptance_number=acceptance_number)
        else:
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
            print("Form errors:", form.errors)
    else:
        form = CarPartsForm(request=request)

    previous_page = request.GET.get('next', 'select_car_for_parts')
    previous_url = reverse(previous_page)
    if previous_page == 'registration_list':
        search_params = request.session.get('search_params', {})
        query_string = '&'.join([f"{k}={v}" for k, v in search_params.items() if v])
        if query_string:
            previous_url += f"?{query_string}"

    return render(request, 'core/add_car_parts_step2.html', {
        'form': form,
        'car': car,
        'previous_url': previous_url,
    })
logger = logging.getLogger(__name__)

@login_required
def add_car_costs_step3(request, acceptance_number):
    sys.stdout.reconfigure(encoding='utf-8')
    logger.debug(f"Starting add_car_costs_step3 for user: {request.user}, acceptance_number: {acceptance_number}")

    car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)

    # چک کردن دسترسی شرکت
    if not request.user.is_superuser and hasattr(car, 'company') and car.company != request.user.userprofile.company:
        logger.warning(f"User {request.user} has no permission for car {acceptance_number}")
        messages.error(request, 'شما اجازه ثبت هزینه برای این خودرو را ندارید.')
        return redirect('registration_list')

    # بررسی وجود هزینه‌های قبلی
    if hasattr(car, 'costs') and car.costs:
        logger.warning(f"Costs already exist for car {acceptance_number}")
        messages.error(request, 'هزینه‌های این خودرو قبلاً ثبت شده است. لطفاً از ویرایش هزینه‌ها استفاده کنید.')
        return redirect('edit_car_costs', acceptance_number=car.acceptance_number)

    # گرفتن مقصد بازگشت از پارامتر next یا سشن (اولویت با GET)
    next_url = request.GET.get('next', request.session.get('return_to', 'registration_list'))
    logger.debug(f"Next URL: {next_url}")

    # گرفتن پارامترهای سرچ از سشن
    search_params = request.session.get('search_params', {})
    logger.debug(f"Search params from session: {search_params}")

    if request.method == 'POST':
        form = CarCostsForm(request.POST, request.FILES, request=request)
        if form.is_valid():
            car_costs = form.save(commit=False)
            car_costs.car = car

            if hasattr(request.user,
                       'userprofile') and request.user.userprofile.first_name and request.user.userprofile.last_name:
                car_costs.recorded_by = f"{request.user.userprofile.first_name} {request.user.userprofile.last_name}"
            else:
                car_costs.recorded_by = request.user.username

            car_costs.company = car.company if hasattr(car, 'company') else None
            if request.user.is_superuser:
                car_costs.company = form.cleaned_data['company'] or car.company

            if not car_costs.company:
                logger.error("No valid company for car or user")
                messages.error(request, "خودرو یا کاربر شرکت معتبری ندارد. لطفاً شرکت را تنظیم کنید.")
                return render(request, 'core/add_car_costs_step3.html', {'form': form, 'car': car, 'next_url': next_url})

            try:
                car_costs.save()
                logger.info(f"Costs saved for car {acceptance_number}")
                messages.success(request, f"هزینه‌های خودرو با شماره پذیرش {car.acceptance_number} با موفقیت ثبت شد.")

                # ساخت URL با توجه به next_url
                redirect_url = reverse(next_url)
                if next_url == 'registration_list' and search_params:
                    query_string = '&'.join([f"{key}={value}" for key, value in search_params.items()])
                    redirect_url += f"?{query_string}"
                logger.debug(f"Redirecting to: {redirect_url}")
                return HttpResponseRedirect(redirect_url)

            except IntegrityError as e:
                logger.error(f"IntegrityError saving costs: {str(e)}")
                messages.error(request, 'خطا: هزینه‌ها برای این خودرو قبلاً ثبت شده است.')
                return render(request, 'core/add_car_costs_step3.html', {'form': form, 'car': car, 'next_url': next_url})
        else:
            logger.debug(f"Form invalid: {form.errors.as_text()}")
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
            return render(request, 'core/add_car_costs_step3.html', {'form': form, 'car': car, 'next_url': next_url})
    else:
        initial = {'company': car.company} if hasattr(car, 'company') else {}
        form = CarCostsForm(request=request, initial=initial)
        logger.debug("Rendering form for GET request")

    # ذخیره مقصد بازگشت توی سشن
    request.session['return_to'] = next_url

    return render(request, 'core/add_car_costs_step3.html', {
        'form': form,
        'car': car,
        'next_url': next_url,
        'search_params': search_params
    })


@login_required
def edit_car_parts(request, acceptance_number):
    sys.stdout.reconfigure(encoding='utf-8')
    car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)

    if not request.user.is_superuser and car.company != request.user.userprofile.company:
        messages.error(request, 'شما اجازه ویرایش قطعات این خودرو را ندارید.')
        return redirect('registration_list')

    parts = car.parts
    if not parts:
        messages.error(request, 'قطعاتی برای ویرایش وجود ندارد.')
        return redirect('registration_list')

    if request.method == 'POST':
        form = CarPartsForm(request.POST, instance=parts, request=request)
        if form.is_valid():
            car_parts = form.save(commit=False)
            changes = {}
            old_instance = CarEntry.objects.get(pk=car.pk).parts
            for field in form.changed_data:
                old_value = getattr(old_instance, field)
                new_value = form.cleaned_data[field]
                if old_value != new_value:
                    changes[field] = {'old': str(old_value), 'new': str(new_value)}

            if request.user.is_superuser:
                car_parts.company = form.cleaned_data['company']
            car_parts.save()

            EditLog.objects.create(
                edit_type='CAR_PARTS',
                car_entry=car,
                edited_by=request.user,
                company=car.company,
                changes=changes if changes else None
            )
            messages.success(request, f"قطعات خودرو با شماره پذیرش {car.acceptance_number} با موفقیت ویرایش شد.")

            # بررسی صفحه بعدی
            next_page = request.GET.get('next', 'registration_list')
            if next_page == 'registration_list':
                search_params = request.session.get('search_params', {})
                query_string = '&'.join([f"{k}={v}" for k, v in search_params.items() if v])
                redirect_url = reverse('registration_list')
                if query_string:
                    redirect_url += f"?{query_string}"
                return redirect(redirect_url)
            return redirect(next_page)
        else:
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
            print("Form errors:", form.errors)
    else:
        form = CarPartsForm(instance=parts, request=request)

    # تنظیم صفحه قبلی برای بازگشت
    previous_page = request.GET.get('next', 'registration_list')  # پیش‌فرض به registration_list
    previous_url = reverse(previous_page)
    if previous_page == 'registration_list':
        search_params = request.session.get('search_params', {})
        query_string = '&'.join([f"{k}={v}" for k, v in search_params.items() if v])
        if query_string:
            previous_url += f"?{query_string}"

    return render(request, 'core/edit_car_parts.html', {
        'form': form,
        'car': car,
        'previous_url': previous_url,
    })

logger = logging.getLogger(__name__)

@login_required
def edit_car_costs(request, acceptance_number, form=None):
    sys.stdout.reconfigure(encoding='utf-8')
    logger.debug(f"Starting edit_car_costs for user: {request.user}, acceptance_number: {acceptance_number}")

    # گرفتن خودرو
    car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)
    logger.debug(f"Car data: {car.__dict__}")

    # چک کردن دسترسی
    if not request.user.is_superuser and car.company != request.user.userprofile.company:
        logger.warning(f"User {request.user} has no permission for car {acceptance_number}")
        messages.error(request, 'شما اجازه ویرایش هزینه برای این خودرو را ندارید.')
        return redirect('registration_list')

    # گرفتن هزینه‌ها
    costs = car.costs
    logger.debug(f"Costs object: {costs}")
    if not costs:
        logger.warning(f"No costs found for car {acceptance_number}")
        messages.error(request, 'هزینه‌های این خودرو هنوز ثبت نشده است.')
        return redirect('add_car_costs_step3', acceptance_number=car.acceptance_number)
    logger.debug(f"Costs data: {costs.__dict__}")

    # گرفتن مقصد بازگشت
    next_url = request.GET.get('next', request.session.get('return_to', 'registration_list'))
    search_params = request.session.get('search_params', {})
    logger.debug(f"Next URL: {next_url}, Search params: {search_params}")

    if request.method == 'POST':
        if form is None:
            form = CarCostsForm(request.POST, request.FILES, instance=costs, request=request)
        logger.debug(f"POST form data: {form.data}")
        if form.is_valid():
            car_costs = form.save(commit=False)
            changes = {}
            old_instance = CarEntry.objects.get(pk=car.pk).costs
            for field in form.changed_data:
                old_value = getattr(old_instance, field)
                new_value = form.cleaned_data[field]
                if old_value != new_value:
                    changes[field] = {'old': str(old_value), 'new': str(new_value)}

            if request.user.is_superuser:
                car_costs.company = form.cleaned_data['company']

            car_costs.save()
            logger.info(f"Costs updated for car {acceptance_number}")

            EditLog.objects.create(
                edit_type='CAR_COSTS',
                car_entry=car,
                edited_by=request.user,
                company=car.company,
                changes=changes if changes else None
            )
            messages.success(request, f"هزینه‌های خودرو با شماره پذیرش {car.acceptance_number} با موفقیت ویرایش شد.")

            redirect_url = reverse(next_url)
            if next_url == 'registration_list' and search_params:
                query_string = '&'.join([f"{key}={value}" for key, value in search_params.items()])
                redirect_url += f"?{query_string}"
            logger.debug(f"Redirecting to: {redirect_url}")
            return redirect(redirect_url)
        else:
            logger.debug(f"Form invalid: {form.errors.as_text()}")
            messages.error(request, 'لطفاً خطاها را برطرف کنید.')
    else:
        if form is None:
            form = CarCostsForm(instance=costs, request=request)
        logger.debug(f"Form initial data: {form.initial}")
        logger.debug(f"Form fields: { {field.name: field.value() for field in form} }")
        logger.debug("Rendering form for GET request")

    return render(request, 'core/edit_car_costs.html', {
        'form': form,
        'car': car,
        'next_url': next_url,
        'search_params': search_params
    })
@login_required
def car_details(request, acceptance_number):
    sys.stdout.reconfigure(encoding='utf-8')
    car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)

    # خروجی PDF
    if 'export_pdf' in request.GET:
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="car_details_{acceptance_number}.pdf"'
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)

        # ثبت فونت فارسی
        font_path = 'static/fonts/Vazirmatn-Regular.ttf'  # مسیر فونت رو چک کن
        pdfmetrics.registerFont(TTFont('Vazir', font_path))
        p.setFont('Vazir', 12)

        # تابع برای آماده‌سازی متن RTL
        def prepare_rtl_text(text):
            reshaped_text = arabic_reshaper.reshape(str(text))
            return get_display(reshaped_text)

        # عرض و ارتفاع صفحه
        page_width = A4[0]  # 595.27 pt
        page_height = A4[1]  # 841.89 pt
        column_width = page_width / 3  # تقریباً 198 pt برای هر ستون
        # تقسیم ارتفاع: اطلاعات خودرو و قطعات و هزینه‌ها هر کدوم 35%، سند 10%
        total_content_height = page_height - 100  # 100 pt برای عنوان
        section_height_main = total_content_height * 0.35  # برای اطلاعات خودرو، قطعات، هزینه‌ها
        section_height_doc = total_content_height * 0.10  # برای سند (فضای کمتر)
        margin = 20  # فاصله از لبه‌ها و بین ستون‌ها

        # مختصات شروع هر ستون در هر بخش (راست‌چین)
        col1_x = page_width - margin
        col2_x = page_width - column_width - margin
        col3_x = page_width - (2 * column_width) - margin

        # عنوان (وسط صفحه)
        p.setFont('Vazir', 16)
        p.setFillColor(colors.black)
        title = prepare_rtl_text(f"جزئیات خودرو: {car.acceptance_number}")
        title_width = p.stringWidth(title, 'Vazir', 16)
        p.drawString((page_width - title_width) / 2, page_height - 50, title)

        # بخش‌ها
        sections = [
            ("اطلاعات خودرو", [
                f"شماره پارکینگ: {car.parking_number or ''}",
                f"تاریخ تحویل: {car.delivery_date or ''}",
                f"نوع خودرو: {car.car_type or ''}",
                f"شماره انتظامی: {car.license_plate or ''}",
                f"نام مالک: {car.owner_name or ''}",
                f"شماره موتور: {car.engine_number or ''}",
                f"شماره شاسی: {car.chassis_number or ''}",
                f"پلاک اتاق: {'دارد' if car.has_cabin_plate else 'ندارد'}",
                f"شماره پذیرش: {car.acceptance_number}",
                f"مسئول پذیرش: {car.accepted_by or ''}",
                f"تاریخ پذیرش: {car.accepted_at or ''}",
                f"ویرایش‌کننده: {car.edited_by or ''}",
                f"نام راننده: {car.driver_name or ''}",
                f"شماره پلاک راننده: {car.driver_license_plate or ''}",
                f"شماره تماس راننده: {car.driver_phone or ''}",
            ], section_height_main),
            ("قطعات", [
                f"درصد اتاق: {car.parts.cabin_percentage or ''}" if hasattr(car,
                                                                            'parts') and car.parts else "قطعات هنوز ثبت نشده است.",
                (f"درب موتور: {'دارد' if car.parts.hood else 'ندارد'}",
                 car.parts.hood if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                              'parts') and car.parts else "",
                (f"رادیاتور: {'دارد' if car.parts.radiator else 'ندارد'}",
                 car.parts.radiator if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                  'parts') and car.parts else "",
                (f"ایسیو: {'دارد' if car.parts.ecu else 'ندارد'}",
                 car.parts.ecu if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                             'parts') and car.parts else "",
                (f"درب جلو و عقب: {'دارد' if car.parts.front_rear_doors else 'ندارد'}",
                 car.parts.front_rear_doors if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                          'parts') and car.parts else "",
                (f"سیم‌کشی: {'دارد' if car.parts.wiring else 'ندارد'}",
                 car.parts.wiring if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                'parts') and car.parts else "",
                (f"کابل باطری: {'دارد' if car.parts.battery_cable else 'ندارد'}",
                 car.parts.battery_cable if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                       'parts') and car.parts else "",
                (f"بخاری: {'دارد' if car.parts.heater else 'ندارد'}",
                 car.parts.heater if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                'parts') and car.parts else "",
                (f"فنر زیر و بند: {'دارد' if car.parts.suspension_spring else 'ندارد'}",
                 car.parts.suspension_spring if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                           'parts') and car.parts else "",
                (f"دینام: {'دارد' if car.parts.alternator else 'ندارد'}",
                 car.parts.alternator if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                    'parts') and car.parts else "",
                (f"موتور برف‌پاک‌کن: {'دارد' if car.parts.wiper_motor else 'ندارد'}",
                 car.parts.wiper_motor if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                     'parts') and car.parts else "",
                (f"رینگ و لاستیک: {'دارد' if car.parts.rims_tires else 'ندارد'}",
                 car.parts.rims_tires if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                    'parts') and car.parts else "",
                (f"کاربراتور: {'دارد' if car.parts.carburetor else 'ندارد'}",
                 car.parts.carburetor if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                    'parts') and car.parts else "",
                (f"استارت: {'دارد' if car.parts.starter else 'ندارد'}",
                 car.parts.starter if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                 'parts') and car.parts else "",
                (f"گیربکس: {'دارد' if car.parts.gearbox else 'ندارد'}",
                 car.parts.gearbox if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                 'parts') and car.parts else "",
                (f"دیفرانسیل: {'دارد' if car.parts.differential else 'ندارد'}",
                 car.parts.differential if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                      'parts') and car.parts else "",
                (f"دلکو و کوئل: {'دارد' if car.parts.distributor_coil else 'ندارد'}",
                 car.parts.distributor_coil if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                          'parts') and car.parts else "",
                (f"سیفون بنزین: {'دارد' if car.parts.fuel_pump else 'ندارد'}",
                 car.parts.fuel_pump if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                   'parts') and car.parts else "",
                (f"صندلی: {'دارد' if car.parts.seats else 'ندارد'}",
                 car.parts.seats if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                               'parts') and car.parts else "",
                (f"دیسک چرخ جلو: {'دارد' if car.parts.front_disc else 'ندارد'}",
                 car.parts.front_disc if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                    'parts') and car.parts else "",
                (f"باطری: {'دارد' if car.parts.battery else 'ندارد'}",
                 car.parts.battery if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                 'parts') and car.parts else "",
                (f"جعبه فرمان: {'دارد' if car.parts.steering_box else 'ندارد'}",
                 car.parts.steering_box if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                      'parts') and car.parts else "",
                (f"سپر عقب و جلو: {'دارد' if car.parts.bumpers else 'ندارد'}",
                 car.parts.bumpers if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                 'parts') and car.parts else "",
                (f"کاسه چرخ: {'دارد' if car.parts.wheel_drum else 'ندارد'}",
                 car.parts.wheel_drum if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                    'parts') and car.parts else "",
                (f"پلوس و گاردان: {'دارد' if car.parts.driveshaft_cv else 'ندارد'}",
                 car.parts.driveshaft_cv if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                       'parts') and car.parts else "",
                (f"ریل سوخت: {'دارد' if car.parts.fuel_rail else 'ندارد'}",
                 car.parts.fuel_rail if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                   'parts') and car.parts else "",
                (f"درب صندوق: {'دارد' if car.parts.trunk_lid else 'ندارد'}",
                 car.parts.trunk_lid if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                   'parts') and car.parts else "",
                (f"سند: {'دارد' if car.parts.documents else 'ندارد'}",
                 car.parts.documents if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                   'parts') and car.parts else "",
                (f"انژکتور و سوزن انژکتور: {'دارد' if car.parts.injector else 'ندارد'}",
                 car.parts.injector if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                  'parts') and car.parts else "",
                (f"شیشه: {'دارد' if car.parts.glass else 'ندارد'}",
                 car.parts.glass if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                               'parts') and car.parts else "",
                (f"کولر: {'دارد' if car.parts.ac else 'ندارد'}",
                 car.parts.ac if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                            'parts') and car.parts else "",
                f"کپسول: {car.parts.get_gas_cylinder_display() if hasattr(car, 'parts') and car.parts else ''}",
                (f"پوستر ترمز: {'دارد' if car.parts.brake_booster else 'ندارد'}",
                 car.parts.brake_booster if hasattr(car, 'parts') and car.parts else False) if hasattr(car,
                                                                                                       'parts') and car.parts else "",
                f"وزن: {car.parts.weight or ''}" if hasattr(car, 'parts') and car.parts else "",
                f"وضعیت پلاک: {car.parts.get_plate_status_display() if hasattr(car, 'parts') and car.parts else ''}",
                f"توضیحات پلاک: {car.parts.plate_description or ''}" if hasattr(car, 'parts') and car.parts else "",
                f"مسئول ثبت: {car.parts.recorded_by.username if car.parts.recorded_by else ''}" if hasattr(car,
                                                                                                           'parts') and car.parts else "",
                f"تاریخ ثبت: {car.parts.recorded_at or ''}" if hasattr(car, 'parts') and car.parts else "",
            ], section_height_main),
            ("هزینه‌ها", [
                f"قیمت روز: {car.costs.daily_price or ''}" if hasattr(car,
                                                                      'costs') and car.costs else "هزینه‌ها هنوز ثبت نشده است.",
                f"خلافی: {car.costs.fine or ''}" if hasattr(car, 'costs') and car.costs else "",
                f"سند خلافی: {car.costs.fine_document.name if car.costs.fine_document else ''}" if hasattr(car,
                                                                                                           'costs') and car.costs else "",
                f"هزینه دادگاهی: {car.costs.court_cost or ''}" if hasattr(car, 'costs') and car.costs else "",
                f"سند دادگاهی: {car.costs.court_document.name if car.costs.court_document else ''}" if hasattr(car,
                                                                                                               'costs') and car.costs else "",
                f"جمع کل: {car.costs.total_cost or ''}" if hasattr(car, 'costs') and car.costs else "",
                f"مسئول ثبت: {car.costs.recorded_by.username if car.costs.recorded_by else ''}" if hasattr(car,
                                                                                                           'costs') and car.costs else "",
                f"تاریخ ثبت: {car.costs.recorded_at or ''}" if hasattr(car, 'costs') and car.costs else "",
            ], section_height_main),
            ("سند", [
                (f"فایل ۱: {car.document.file1.name if car.document.file1 else ''}",
                 bool(car.document.file1)) if hasattr(car, 'document') and car.document else "سند هنوز ثبت نشده است.",
                (f"فایل ۲: {car.document.file2.name if car.document.file2 else ''}",
                 bool(car.document.file2)) if hasattr(car, 'document') and car.document else "",
                (f"فایل ۳: {car.document.file3.name if car.document.file3 else ''}",
                 bool(car.document.file3)) if hasattr(car, 'document') and car.document else "",
                f"ایجاد شده توسط: {car.document.created_by.username if car.document.created_by else ''}" if hasattr(car,'document') and car.document else "",
                f"تاریخ ایجاد: { car.document.created_at or ''}" if hasattr(car, 'document') and car.document else "",
            ], section_height_doc),
        ]

        # نمایش هر بخش با 3 ستون
        y_start = page_height - 100
        for section_idx, (section_title, section_data, section_height) in enumerate(sections):
            y = y_start - (sum(sections[i][2] for i in range(section_idx)))  # موقعیت عمودی هر بخش
            p.setFont('Vazir', 14)
            p.setFillColor(colors.black)
            rtl_title = prepare_rtl_text(section_title)
            title_width = p.stringWidth(rtl_title, 'Vazir', 14)
            p.drawString((page_width - title_width) / 2, y, rtl_title)
            y -= 20

            # حذف خطوط خالی و تقسیم داده‌ها به 3 ستون
            filtered_data = [item if isinstance(item, str) else item[0] for item in section_data if item]
            items_per_column = len(filtered_data) // 3 + (len(filtered_data) % 3 > 0)
            col1_data = filtered_data[:items_per_column]
            col2_data = filtered_data[items_per_column:2 * items_per_column]
            col3_data = filtered_data[2 * items_per_column:]

            # تنظیمات برای رنگ (آبی برای قطعات، سبز برای فایل‌ها)
            boolean_items = {item[0]: item[1] for item in section_data if isinstance(item, tuple)}

            # ستون اول
            p.setFont('Vazir', 10)
            y_col = y
            for line in col1_data:
                rtl_line = prepare_rtl_text(line)
                line_width = p.stringWidth(rtl_line, 'Vazir', 10)
                if line in boolean_items:
                    if section_title == "سند" and boolean_items[line]:
                        p.setFillColor(colors.green)  # سبز برای فایل‌های آپلود شده
                    elif section_title == "قطعات" and boolean_items[line]:
                        p.setFillColor(colors.blue)  # آبی برای قطعات True
                    else:
                        p.setFillColor(colors.black)
                else:
                    p.setFillColor(colors.black)
                p.drawString(col1_x - line_width, y_col, rtl_line)
                y_col -= 15

            # ستون دوم
            y_col = y
            for line in col2_data:
                rtl_line = prepare_rtl_text(line)
                line_width = p.stringWidth(rtl_line, 'Vazir', 10)
                if line in boolean_items:
                    if section_title == "سند" and boolean_items[line]:
                        p.setFillColor(colors.green)
                    elif section_title == "قطعات" and boolean_items[line]:
                        p.setFillColor(colors.blue)
                    else:
                        p.setFillColor(colors.black)
                else:
                    p.setFillColor(colors.black)
                p.drawString(col2_x - line_width, y_col, rtl_line)
                y_col -= 15

            # ستون سوم
            y_col = y
            for line in col3_data:
                rtl_line = prepare_rtl_text(line)
                line_width = p.stringWidth(rtl_line, 'Vazir', 10)
                if line in boolean_items:
                    if section_title == "سند" and boolean_items[line]:
                        p.setFillColor(colors.green)
                    elif section_title == "قطعات" and boolean_items[line]:
                        p.setFillColor(colors.blue)
                    else:
                        p.setFillColor(colors.black)
                else:
                    p.setFillColor(colors.black)
                p.drawString(col3_x - line_width, y_col, rtl_line)
                y_col -= 15

        p.showPage()
        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    return render(request, 'core/car_details.html', {'car': car})
# from django.contrib.auth.decorators import login_required
# from django.shortcuts import render, redirect
# from django.contrib import messages
# from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
# from django.db.models import Q
# from .models import CarEntry, CarParts
# from openpyxl import Workbook
# from django.http import HttpResponse
# import re
# import jdatetime
# from jdatetime import datetime as jdatetime_datetime

# تنظیم لاگ برای دیباگ
logger = logging.getLogger(__name__)

logger = logging.getLogger(__name__)

logger = logging.getLogger(__name__)

logger = logging.getLogger(__name__)

@login_required
def registration_list(request):
    sys.stdout.reconfigure(encoding='utf-8')
    logger.debug(f"Starting registration_list for user: {request.user}, method: {request.method}")

    # چک کردن شرکت و فقط فعال‌ها
    if not request.user.is_superuser:
        try:
            if not hasattr(request.user, 'userprofile') or not request.user.userprofile.company:
                logger.warning("User has no profile or company, redirecting to company_select")
                messages.error(request, 'شما به هیچ شرکتی متصل نیستید.')
                return redirect('company_select')
            company = request.user.userprofile.company
            cars = CarEntry.objects.filter(company=company).order_by('-delivery_date_gregorian')
            logger.debug(f"Company: {company}, Initial queryset count: {cars.count()}")
        except Exception as e:
            logger.error(f"Error in company check: {str(e)}")
            messages.error(request, f'خطا در دسترسی به شرکت: {str(e)}')
            return redirect('company_select')
    else:
        company = None
        try:
            cars = CarEntry.objects.all.order_by('-delivery_date_gregorian')

            logger.debug(f"Superuser, Initial queryset count: {cars.count()}")
        except Exception as e:
            logger.error(f"Error fetching cars for superuser: {str(e)}")
            messages.error(request, f'خطا در بارگذاری داده‌ها: {str(e)}')
            cars = CarEntry.objects.none()

    # گرفتن مقادیر فیلترها
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    license_plate = request.GET.get('license_plate', '').strip()
    parking_number = request.GET.get('parking_number', '').strip()
    engine_number = request.GET.get('engine_number', '').strip()
    has_parts = request.GET.get('has_parts', '')
    has_costs = request.GET.get('has_costs', '')
    has_document = request.GET.get('has_document', '')
    chassis_number = request.GET.get('chassis_number', '').strip()
    has_complaints = request.GET.get('has_complaints', '')
    # has_serial_number = request.GET.get('has_serial_number', '')
    physical_readiness = request.GET.get('physical_readiness', '')
    has_all_images = request.GET.get('has_all_images', '')
    page_number = request.GET.get('page', '')

    logger.debug(f"Search params: {start_date=}, {end_date=}, {license_plate=}, {page_number=}")

    # اعتبارسنجی تاریخ‌ها
    date_pattern = r'^\d{4}/\d{2}/\d{2}$'
    if start_date and not re.match(date_pattern, start_date):
        messages.error(request, 'فرمت "از تاریخ" اشتباه است. لطفاً از فرمت 1403/12/15 استفاده کنید.')
        start_date = ''
    if end_date and not re.match(date_pattern, end_date):
        messages.error(request, 'فرمت "تا تاریخ" اشتباه است. لطفاً از فرمت 1403/12/15 استفاده کنید.')
        end_date = ''
    if start_date and end_date:
        try:
            start_jdate = jdatetime_datetime.strptime(start_date, '%Y/%m/%d')
            end_jdate = jdatetime_datetime.strptime(end_date, '%Y/%m/%d')
            if start_jdate > end_jdate:
                messages.error(request, '"از تاریخ" نمی‌تواند بزرگ‌تر از "تا تاریخ" باشد.')
                start_date = ''
                end_date = ''
        except ValueError:
            messages.error(request, 'تاریخ‌ها نامعتبر هستند.')
            start_date = ''
            end_date = ''

    # اعمال فیلترها با مدیریت خطا
    try:
        if license_plate and len(license_plate) <= 30:
            cars = cars.filter(license_plate__icontains=license_plate)
        elif license_plate:
            messages.error(request, 'شماره انتظامی نمی‌تواند بیشتر از 30 کاراکتر باشد.')

        if parking_number and len(parking_number) <= 20:
            cars = cars.filter(parking_number__icontains=parking_number)
        elif parking_number:
            messages.error(request, 'شماره پارکینگ نمی‌تواند بیشتر از 20 کاراکتر باشد.')

        if engine_number and len(engine_number) <= 50:
            cars = cars.filter(engine_number__icontains=engine_number)
        elif engine_number:
            messages.error(request, 'شماره موتور نمی‌تواند بیشتر از 50 کاراکتر باشد.')

        if chassis_number and len(chassis_number) <= 50:
            cars = cars.filter(chassis_number__icontains=chassis_number)
        elif chassis_number:
            messages.error(request, 'شماره شاسی نمی‌تواند بیشتر از 30 کاراکتر باشد.')

        if start_date:
            try:
                start_date_gregorian = jdatetime_datetime.strptime(start_date, '%Y/%m/%d').togregorian()
                cars = cars.filter(delivery_date_gregorian__gte=start_date_gregorian)
            except ValueError:
                messages.error(request, 'تاریخ شروع نامعتبر است.')

        if end_date:
            try:
                end_date_gregorian = jdatetime_datetime.strptime(end_date, '%Y/%m/%d').togregorian()
                cars = cars.filter(delivery_date_gregorian__lte=end_date_gregorian)
            except ValueError:
                messages.error(request, 'تاریخ پایان نامعتبر است.')

        if has_parts in ('yes', 'no'):
            cars = cars.filter(parts__isnull=(has_parts == 'no'))
        if has_costs in ('yes', 'no'):
            cars = cars.filter(costs__isnull=(has_costs == 'no'))
        if has_document in ('yes', 'no'):
            if has_document == 'yes':
                # سندهای کامل
                complete_docs = Document.objects.filter(
                    Q(is_deceased=True,
                      car_document__isnull=False,
                      vekalat__isnull=False,
                      deceased_doc1__isnull=False,
                      deceased_doc2__isnull=False) |
                    Q(is_deceased=False,
                      car_document__isnull=False,
                      vekalat__isnull=False)
                ).exclude(
                    Q(car_document='') |
                    Q(vekalat='') |
                    Q(is_deceased=True) & (Q(deceased_doc1='') | Q(deceased_doc2=''))
                )
                print(f"Complete docs count: {complete_docs.count()}")
                print(
                    f"Complete docs: {list(complete_docs.values('engine_number', 'is_deceased', 'car_document', 'vekalat', 'deceased_doc1', 'deceased_doc2'))}")
                cars = cars.filter(document__in=complete_docs).distinct()
                print(f"Filtered cars count (yes): {cars.count()}")
                print(
                    f"Filtered cars (yes): {list(cars.values('acceptance_number', 'engine_number', 'chassis_number'))}")
            elif has_document == 'no':
                # خودروهای بدون سند یا با سند ناقص
                incomplete_docs = Document.objects.filter(
                    # سندهای فوتی ناقص (حداقل یکی از ۴ فیلد خالی یا null)
                    Q(is_deceased=True) & (
                            Q(car_document__isnull=True) | Q(car_document='') |
                            Q(vekalat__isnull=True) | Q(vekalat='') |
                            Q(deceased_doc1__isnull=True) | Q(deceased_doc1='') |
                            Q(deceased_doc2__isnull=True) | Q(deceased_doc2='')
                    ) |
                    # سندهای غیرفوتی ناقص (حداقل یکی از ۲ فیلد خالی یا null)
                    Q(is_deceased=False) & (
                            Q(car_document__isnull=True) | Q(car_document='') |
                            Q(vekalat__isnull=True) | Q(vekalat='')
                    )
                )
                print(f"Incomplete docs count: {incomplete_docs.count()}")
                print(
                    f"Incomplete docs: {list(incomplete_docs.values('engine_number', 'is_deceased', 'car_document', 'vekalat', 'deceased_doc1', 'deceased_doc2'))}")
                cars = cars.filter(
                    Q(document__isnull=True) |  # بدون سند
                    Q(document__in=incomplete_docs)  # یا با سند ناقص
                ).distinct()
                print(f"Filtered cars count (no): {cars.count()}")
                print(
                    f"Filtered cars (no): {list(cars.values('acceptance_number', 'engine_number', 'chassis_number'))}")
        if has_complaints in ('yes', 'no'):
            cars = cars.annotate(has_complaint=Count('complaint')).filter(has_complaint__gt=0 if has_complaints == 'yes' else 0)
        # if has_serial_number in ('yes', 'no'):
        #     cars = cars.filter(serial_number__isnull=(has_serial_number == 'no'))
        if physical_readiness in ('yes', 'no'):
            cars = cars.filter(Physical_readiness=(physical_readiness == 'yes'))
        if has_all_images in ('yes', 'no'):
            if has_all_images == 'yes':
                cars = cars.filter(

                    front_image__isnull=False,
                    rear_image__isnull=False,



                    # chassis_image__isnull=False,
                    # engine_image__isnull=False
                )
            else:
                cars = cars.filter(
                    Q(front_image__isnull=True) |
                    Q(rear_image__isnull=True)

                    # Q(chassis_image__isnull=True) |
                    # Q(engine_image__isnull=True)
                )

        logger.debug(f"After filters, queryset count: {cars.count()}")

    except Exception as e:
        logger.error(f"Error applying filters: {str(e)}")
        messages.error(request, f'خطا در اعمال فیلترها: {str(e)}')

    # ذخیره پارامترهای سرچ توی سشن
    logger.debug("Saving search params to session")
    search_params = {
        'start_date': start_date,
        'end_date': end_date,
        'license_plate': license_plate,
        'parking_number': parking_number,
        'engine_number': engine_number,
        'has_parts': has_parts,
        'has_costs': has_costs,
        'has_document': has_document,
        'chassis_number': chassis_number,
        'has_complaints': has_complaints,
        # 'has_serial_number': has_serial_number,
        'physical_readiness': physical_readiness,
        'has_all_images': has_all_images,
        'page': page_number
    }
    try:
        request.session['search_params'] = {k: v for k, v in search_params.items() if v}
        logger.debug("Search params saved to session")
    except Exception as e:
        logger.error(f"Error saving session: {str(e)}")
        messages.error(request, f'خطا در ذخیره سشن: {str(e)}')

    # ساخت کوئری استرینگ برای ریدایرکت‌ها
    logger.debug("Building query string")
    query_string = ''
    try:
        if search_params:
            query_string = '&'.join([f"{key}={value}" for key, value in search_params.items() if value])
        logger.debug(f"Query string built: {query_string}")
    except Exception as e:
        logger.error(f"Error building query string: {str(e)}")
        messages.error(request, f'خطا در ساخت کوئری: {str(e)}')

    logger.debug("Checking POST and export conditions")
    # مدیریت درخواست‌های POST
    if request.method == 'POST':
        logger.debug("Handling POST request")
        selected_car = request.POST.get('selected_car')
        if not selected_car:
            messages.error(request, 'لطفاً یک ماشین را انتخاب کنید.')
            redirect_url = reverse('registration_list')
            if query_string:
                redirect_url += f"?{query_string}"
            return redirect(redirect_url)

        try:
            car = CarEntry.objects.get(acceptance_number=selected_car)
            base_redirect_url = None
            if 'edit_entry' in request.POST:
                base_redirect_url = reverse('edit_car_entry', kwargs={'acceptance_number': selected_car})
            elif 'edit_parts' in request.POST:
                if hasattr(car, 'parts') and car.parts:
                    base_redirect_url = reverse('edit_car_parts', kwargs={'acceptance_number': selected_car})
                else:
                    messages.error(request, 'قطعات برای این ماشین ثبت نشده است.')
            elif 'add_parts' in request.POST:
                if not (hasattr(car, 'parts') and car.parts):
                    base_redirect_url = reverse('add_car_parts_step2', kwargs={'acceptance_number': selected_car})
                else:
                    messages.error(request, 'قطعات برای این ماشین قبلاً ثبت شده است.')
            elif 'edit_costs' in request.POST:
                if hasattr(car, 'costs') and car.costs:
                    base_redirect_url = reverse('edit_car_costs', kwargs={'acceptance_number': selected_car})
                else:
                    messages.error(request, 'هزینه‌ها برای این ماشین ثبت نشده است.')
            elif 'add_costs' in request.POST:
                if not (hasattr(car, 'costs') and car.costs):
                    base_redirect_url = reverse('add_car_costs_step3', kwargs={'acceptance_number': selected_car})
                else:
                    messages.error(request, 'هزینه‌ها برای این ماشین قبلاً ثبت شده است.')
            elif 'view_details' in request.POST:
                base_redirect_url = reverse('car_details', kwargs={'acceptance_number': selected_car})
            elif 'finalize' in request.POST:
                base_redirect_url = reverse('finalize_registration', kwargs={'acceptance_number': selected_car})
            elif 'add_document' in request.POST:
                if not (hasattr(car, 'document') and car.document):
                    request.session['selected_car'] = selected_car
                    base_redirect_url = reverse('add_document')
                else:
                    messages.error(request, 'سند برای این ماشین قبلاً ثبت شده است.')
            elif 'edit_document' in request.POST:
                if hasattr(car, 'document') and car.document:
                    request.session['selected_car'] = selected_car
                    base_redirect_url = reverse('edit_document', kwargs={'pk': car.document.pk})
                else:
                    messages.error(request, 'سندی برای این ماشین ثبت نشده است.')
            elif 'complaint_action' in request.POST:
                has_complaints = Complaint.objects.filter(car=car).exists()
                if has_complaints:
                    base_redirect_url = reverse('manage_complaints', kwargs={'acceptance_number': selected_car})
                else:
                    base_redirect_url = reverse('add_complaint', kwargs={'acceptance_number': selected_car})
            elif 'serial_action' in request.POST:
                try:
                    car = CarEntry.objects.get(acceptance_number=selected_car)
                    if Serialnumber.objects.filter(car=car).exists():
                        # اگر شماره سریال وجود دارد، به حالت ویرایش می‌رویم
                        base_redirect_url = reverse('update_serial_number',
                                                    kwargs={'acceptance_number': selected_car}) + '?mode=edit'
                    else:
                        # اگر شماره سریال وجود ندارد، به حالت افزودن می‌رویم
                        base_redirect_url = reverse('update_serial_number',
                                                    kwargs={'acceptance_number': selected_car}) + '?mode=add'
                    redirect_url = f"{base_redirect_url}&next=registration_list"
                    if query_string:
                        redirect_url += f"&{query_string}"
                    logger.debug(f"Redirecting to: {redirect_url}")
                    return redirect(redirect_url)
                except CarEntry.DoesNotExist:
                    messages.error(request, 'ماشین انتخاب‌شده وجود ندارد.')


            if base_redirect_url:
                redirect_url = f"{base_redirect_url}?next=registration_list"
                if query_string:
                    redirect_url += f"&{query_string}"
                logger.debug(f"Redirecting to: {redirect_url}")
                return redirect(redirect_url)

        except CarEntry.DoesNotExist:
            messages.error(request, 'ماشین انتخاب‌شده وجود ندارد.')

        # بازگشت پیش‌فرض برای POST
        redirect_url = reverse('registration_list')
        if query_string:
            redirect_url += f"?{query_string}"
        return redirect(redirect_url)

    # خروجی اکسل
    if request.GET.get('export') == 'excel':
        logger.debug("Generating Excel export")
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "لیست ثبت"
            headers = [
                'شماره پذیرش', 'شماره پارکینگ', 'تاریخ تحویل', 'نوع خودرو', 'شماره انتظامی',
                'نام مالک', 'شماره موتور', 'شماره شاسی', 'پلاک اتاق', 'مسئول پذیرش',
                'تاریخ پذیرش', 'نام راننده', 'شماره پلاک راننده', 'شماره تماس راننده',
                # قطعات
                'درصد اتاق', 'درب موتور', 'رادیاتور', 'ایسیو', 'درب جلو و عقب', 'سیم‌کشی',
                'کابل باطری', 'بخاری', 'فنر زیر و بند', 'دینام', 'موتور برف‌پاک‌کن', 'رینگ و لاستیک',
                'کاربراتور', 'استارت', 'گیربکس', 'دیفرانسیل', 'دلکو و کوئل', 'سیفون بنزین', 'صندلی',
                'دیسک چرخ جلو', 'باطری', 'جعبه فرمان', 'سپر عقب و جلو', 'کاسه چرخ', 'پلوس و گاردان',
                'ریل سوخت', 'درب صندوق', 'سند', 'انژکتور و سوزن انژکتور', 'شیشه', 'کولر', 'پوستر ترمز',
                # هزینه‌ها
                'قیمت روز', 'خلافی', 'هزینه دادگاهی', 'هزینه حمل', 'هزینه وکالت', 'جمع کل',
                'مسئول ثبت هزینه‌ها', 'تاریخ ثبت هزینه‌ها'
            ]
            ws.append(headers)

            for car in cars:
                parts_data = car.parts if hasattr(car, 'parts') and car.parts else None
                costs_data = car.costs if hasattr(car, 'costs') and car.costs else None

                row = [
                    car.acceptance_number,
                    car.parking_number,
                    car.delivery_date,
                    car.car_type,
                    car.license_plate,
                    car.owner_name,
                    car.engine_number,
                    car.chassis_number,
                    'دارد' if car.has_cabin_plate else 'ندارد',
                    car.accepted_by if car.accepted_by else '',
                    car.accepted_at,
                    car.driver_name,
                    car.driver_license_plate,
                    car.driver_phone,
                    parts_data.cabin_percentage if parts_data else '',
                    'دارد' if parts_data and parts_data.hood else 'ندارد',
                    'دارد' if parts_data and parts_data.radiator else 'ندارد',
                    'دارد' if parts_data and parts_data.ecu else 'ندارد',
                    'دارد' if parts_data and parts_data.front_rear_doors else 'ندارد',
                    'دارد' if parts_data and parts_data.wiring else 'ندارد',
                    'دارد' if parts_data and parts_data.battery_cable else 'ندارد',
                    'دارد' if parts_data and parts_data.heater else 'ندارد',
                    'دارد' if parts_data and parts_data.suspension_spring else 'ندارد',
                    'دارد' if parts_data and parts_data.alternator else 'ندارد',
                    'دارد' if parts_data and parts_data.wiper_motor else 'ندارد',
                    'دارد' if parts_data and parts_data.rims_tires else 'ندارد',
                    'دارد' if parts_data and parts_data.carburetor else 'ندارد',
                    'دارد' if parts_data and parts_data.starter else 'ندارد',
                    'دارد' if parts_data and parts_data.gearbox else 'ندارد',
                    'دارد' if parts_data and parts_data.differential else 'ندارد',
                    'دارد' if parts_data and parts_data.distributor_coil else 'ندارد',
                    'دارد' if parts_data and parts_data.fuel_pump else 'ندارد',
                    'دارد' if parts_data and parts_data.seats else 'ندارد',
                    'دارد' if parts_data and parts_data.front_disc else 'ندارد',
                    'دارد' if parts_data and parts_data.battery else 'ندارد',
                    'دارد' if parts_data and parts_data.steering_box else 'ندارد',
                    'دارد' if parts_data and parts_data.bumpers else 'ندارد',
                    'دارد' if parts_data and parts_data.wheel_drum else 'ندارد',
                    'دارد' if parts_data and parts_data.driveshaft_cv else 'ندارد',
                    'دارد' if parts_data and parts_data.fuel_rail else 'ندارد',
                    'دارد' if parts_data and parts_data.trunk_lid else 'ندارد',
                    'دارد' if parts_data and parts_data.documents else 'ندارد',
                    'دارد' if parts_data and parts_data.injector else 'ندارد',
                    'دارد' if parts_data and parts_data.glass else 'ندارد',
                    'دارد' if parts_data and parts_data.ac else 'ندارد',
                    'دارد' if parts_data and parts_data.brake_booster else 'ندارد',
                    costs_data.daily_price if costs_data else '',
                    costs_data.fine if costs_data else '',
                    costs_data.court_cost if costs_data else '',
                    costs_data.transport_cost if costs_data else '',
                    costs_data.proxy_cost if costs_data else '',
                    costs_data.total_cost if costs_data else '',
                    costs_data.recorded_by.username if costs_data and costs_data.recorded_by else '',
                    costs_data.recorded_at if costs_data else ''
                ]
                ws.append(row)

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=5, max_col=5):  # ستون ۵ (E)
                for cell in row:
                    cell.alignment = Alignment(horizontal='right', vertical='center')

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="registration_list.xlsx"'
            wb.save(response)
            logger.debug("Returning Excel response")
            return response
        except Exception as e:
            logger.error(f"Error generating Excel: {str(e)}")
            messages.error(request, f'خطا در تولید اکسل: {str(e)}')
            # ادامه به صفحه‌بندی و رندر به جای توقف

    # صفحه‌بندی
    logger.debug("Starting pagination")
    try:
        logger.debug(f"Creating paginator with cars: {cars}")
        paginator = Paginator(cars, 10)
        logger.debug(f"Paginator created with {paginator.count} items")
        page_cars = paginator.page(page_number if page_number else 1)
        logger.debug(f"Page retrieved: {page_cars.number}")
    except (PageNotAnInteger, ValueError):
        logger.warning("Invalid page number, defaulting to page 1")
        page_cars = paginator.page(1)
    except EmptyPage:
        logger.warning("Empty page, defaulting to last page")
        page_cars = paginator.page(paginator.num_pages)
    except Exception as e:
        logger.error(f"Pagination error: {str(e)}")
        messages.error(request, f'خطا در صفحه‌بندی: {str(e)}')
        page_cars = []

    # آماده‌سازی برای رندر
    logger.debug(f"Preparing context with {len(page_cars)} items")
    context = {
        'cars': page_cars,
    }
    logger.debug("Rendering registration_list.html")
    response = render(request, 'core/registration_list.html', context)
    logger.debug("Render complete")
    return response

# @login_required
# def check_status(request, acceptance_number):
#     car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)
#     return JsonResponse({
#         'has_parts': hasattr(car, 'parts') and car.parts is not None,
#         'has_costs': hasattr(car, 'costs') and car.costs is not None,
#         'has_document': hasattr(car, 'document') and car.document is not None,
#     })
@login_required
@require_GET
def check_status(request, acceptance_number):
    car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)

    # چک کردن دسترسی شرکت
    if not request.user.is_superuser and car.company != request.user.userprofile.company:
        return JsonResponse({'error': 'شما اجازه دسترسی به این خودرو را ندارید.'}, status=403)

    # چک کردن وجود روابط
    has_parts = hasattr(car, 'parts') and car.parts is not None
    has_costs = hasattr(car, 'costs') and car.costs is not None
    has_document = hasattr(car, 'document') and car.document is not None
    has_complaints = Complaint.objects.filter(car=car).exists()  # چک کردن وجود شکایت
    # has_serial_number = car.serial_number is not None  # چک کردن null بودن

    return JsonResponse({
        'has_parts': has_parts,
        'has_costs': has_costs,
        'has_document': has_document,
        'has_complaints': has_complaints,
        # 'has_serial_number': has_serial_number  # True یا False برمی‌گردونه
    })
@login_required
def finalize_registration(request, acceptance_number):
    car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)

    # اگر ثبت نهایی وجود داشته باشد، برای ویرایش استفاده می‌شود
    final_reg = getattr(car, 'final_registration', None)

    # گرفتن آدرس بازگشت از سشن
    next_url = request.session.get('return_to', reverse('second_inspection_list'))

    if request.method == 'POST':
        form = FinalRegistrationForm(request.POST, request.FILES, instance=final_reg, request=request)
        if form.is_valid():
            final_reg = form.save(commit=False)
            final_reg.car = car
            final_reg.company = car.company  # گرفتن company از CarEntry
            final_reg.save()
            # غیرفعال کردن CarEntry فقط در ثبت جدید
            if not form.instance.pk:  # اگر ثبت جدید است
                car.is_active = False
                car.save(update_fields=['is_active'])
            messages.success(request, f'خودرو {car.acceptance_number} با موفقیت {"به‌روزرسانی" if form.instance.pk else "ثبت نهایی"} شد.')
            return redirect(next_url)  # ریدایرکت به صفحه قبلی
    else:
        form = FinalRegistrationForm(instance=final_reg, request=request)

    return render(request, 'core/finalize_registration.html', {
        'form': form,
        'car': car,
        'next_url': next_url,  # پاس دادن next_url به قالب
    })


# @login_required
# def add_or_edit_first_inspection(request, acceptance_number, inspection_id=None):
#     sys.stdout.reconfigure(encoding='utf-8')
#
#     # دریافت خودرو
#     try:
#         car = CarEntry.objects.get(acceptance_number=acceptance_number)
#     except CarEntry.DoesNotExist:
#         messages.error(request, f"خودرو با شماره پذیرش {acceptance_number} یافت نشد.")
#         return redirect('home')
#
#     # بررسی دسترسی کاربر
#     if not request.user.is_superuser and car.company != request.user.userprofile.company:
#         messages.error(request, "شما اجازه دسترسی به این خودرو را ندارید.")
#         return redirect('home')
#
#     # اگه inspection_id باشه، ویرایش انجام می‌شه
#     if inspection_id:
#         try:
#             first_inspection = FirstInspection.objects.get(id=inspection_id, car=car)
#         except FirstInspection.DoesNotExist:
#             messages.error(request, "بازدید اول موردنظر یافت نشد.")
#             return redirect('first_inspection_list')
#     else:
#         first_inspection = None
#
#     if request.method == 'POST':
#         form = FirstInspectionForm(request.POST, request.FILES, instance=first_inspection, request=request)
#         if form.is_valid():
#             first_inspection = form.save(commit=False)
#             first_inspection.car = car
#             first_inspection.company = car.company
#
#
#             # مدیریت دکمه‌ها
#             if 'save' in request.POST:
#                 first_inspection.accept = True
#                 messages.success(request,
#                                  f"بازدید اول برای خودرو {car.acceptance_number} با موفقیت {'ویرایش و تأیید' if inspection_id else 'ثبت و تأیید'} شد.")
#             elif 'exit' in request.POST:
#                 first_inspection.exit = True
#                 first_inspection.comment_exit = request.POST.get('comment_exit', '')
#                 if not first_inspection.comment_exit:
#                     messages.error(request, "لطفاً علت خروج را وارد کنید.")
#                     return render(request, 'core/add_first_inspection.html',
#                                   {'form': form, 'car': car, 'inspection': first_inspection})
#                 messages.success(request,
#                                  f"خودرو {car.acceptance_number} از چرخه خارج {'و ویرایش' if inspection_id else ''} شد.")
#             else:  # در لیست بماند
#                 messages.success(request,
#                                  f"بازدید اول برای خودرو {car.acceptance_number} با موفقیت {'ویرایش' if inspection_id else 'ثبت'} شد و در لیست باقی ماند.")
#
#             first_inspection.save()
#             return redirect('first_inspection_list')
#         else:
#             messages.error(request, "لطفاً خطاها را برطرف کنید.")
#             print("Form errors:", form.errors)
#     else:
#         form = FirstInspectionForm(instance=first_inspection, request=request)
#
#     return render(request, 'core/add_first_inspection.html', {
#         'form': form,
#         'car': car,
#         'inspection': first_inspection  # برای تشخیص حالت ویرایش توی قالب
#     })
#
#
# @login_required
# def manage_first_inspection(request, acceptance_number, inspection_id=None):
#     sys.stdout.reconfigure(encoding='utf-8')
#
#     # دریافت خودرو
#     car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)
#
#     # گرفتن شرکت کاربر و فیلتر کردن بازدیدها
#     try:
#         company = request.user.userprofile.company
#         if not company:
#             messages.error(request, "شما به شرکتی متصل نیستید.")
#             return redirect('company_select')
#         if not request.user.is_superuser and car.company != company:
#             messages.error(request, "شما اجازه دسترسی به این خودرو را ندارید.")
#             return redirect('home')
#         first_inspections = FirstInspection.objects.filter(car=car, company=company).order_by(
#             '-inspection_date_gregorian')
#     except AttributeError:
#         messages.error(request, "پروفایل شما کامل نیست.")
#         return redirect('company_select')
#
#     # مدیریت ثبت یا ویرایش
#     if inspection_id:
#         first_inspection = get_object_or_404(FirstInspection, id=inspection_id, car=car, company=company)
#     else:
#         first_inspection = None
#
#     if request.method == 'POST':
#         form = FirstInspectionForm(request.POST, request.FILES, instance=first_inspection, request=request)
#         if form.is_valid():
#             first_inspection = form.save(commit=False)
#             first_inspection.car = car
#             first_inspection.company = company
#
#
#             # مدیریت دکمه‌ها
#             if 'save' in request.POST:
#                 first_inspection.accept = True
#                 messages.success(request,
#                                  f"بازدید اول برای خودرو {car.acceptance_number} با موفقیت {'ویرایش و تأیید' if inspection_id else 'ثبت و تأیید'} شد.")
#             elif 'exit' in request.POST:
#                 first_inspection.exit = True
#                 first_inspection.comment_exit = request.POST.get('comment_exit', '')
#                 if not first_inspection.comment_exit:
#                     messages.error(request, "لطفاً علت خروج را وارد کنید.")
#                     return render(request, 'core/manage_first_inspection.html', {
#                         'form': form, 'car': car, 'first_inspections': first_inspections, 'inspection': first_inspection
#                     })
#                 messages.success(request,
#                                  f"خودرو {car.acceptance_number} از چرخه خارج {'و ویرایش' if inspection_id else ''} شد.")
#             else:  # در لیست بماند
#                 messages.success(request,
#                                  f"بازدید اول برای خودرو {car.acceptance_number} با موفقیت {'ویرایش' if inspection_id else 'ثبت'} شد و در لیست باقی ماند.")
#
#             first_inspection.save()
#             return redirect('manage_first_inspection', acceptance_number=acceptance_number)
#         else:
#             messages.error(request, "لطفاً خطاها را برطرف کنید.")
#             print("Form errors:", form.errors)
#     else:
#         form = FirstInspectionForm(instance=first_inspection, request=request)
#
#     next_url = request.GET.get('next', request.session.get('return_to', 'first_inspection_list'))
#     search_params = request.session.get('search_params', {})
#
#     return render(request, 'core/manage_first_inspection.html', {
#         'form': form,
#         'car': car,
#         'first_inspections': first_inspections,
#         'inspection': first_inspection,
#         'next_url': next_url,
#         'search_params': search_params
#     })


@login_required
def first_inspection_list(request):
    sys.stdout.reconfigure(encoding='utf-8')

    # ذخیره پارامترهای جستجو توی سشن
    if request.method == 'GET' and any(key in request.GET for key in ['q', 'start_date', 'end_date', 'acceptance_number', 'has_inspection', 'engine_number', 'chassis_number']):
        request.session['search_params'] = {
            'q': request.GET.get('q', ''),
            'start_date': request.GET.get('start_date', ''),
            'end_date': request.GET.get('end_date', ''),
            'acceptance_number': request.GET.get('acceptance_number', ''),
            'has_inspection': request.GET.get('has_inspection', ''),
            'engine_number': request.GET.get('engine_number', ''),
            'chassis_number': request.GET.get('chassis_number', '')
        }
    elif 'reset_search' in request.GET:
        request.session['search_params'] = {}

    search_params = request.session.get('search_params', {})

    # فیلتر اولیه خودروها بر اساس شرکت کاربر
    if request.user.is_superuser:
        cars = CarEntry.objects.all()
    else:
        try:
            company = request.user.userprofile.company
            if not company:
                messages.error(request, "شما به شرکتی متصل نیستید.")
                return redirect('company_select')
            cars = CarEntry.objects.filter(company=company)
        except AttributeError:
            messages.error(request, "پروفایل شما کامل نیست.")
            return redirect('company_select')

    # اعمال شروط اولیه
    cars = cars.filter(
        documents__isnull=False,
        Physical_readiness=True
    ).distinct()

    # سندهای کامل
    complete_docs = Document.objects.filter(
        Q(is_deceased=True,
          car_document__isnull=False,
          vekalat__isnull=False,
          deceased_doc1__isnull=False,
          deceased_doc2__isnull=False) |
        Q(is_deceased=False,
          car_document__isnull=False,
          vekalat__isnull=False)
    ).exclude(
        Q(car_document='') |
        Q(vekalat='') |
        Q(is_deceased=True) & (Q(deceased_doc1='') | Q(deceased_doc2=''))
    )

    # فیلتر خودروها بر اساس سندهای کامل
    cars = cars.filter(document__in=complete_docs).distinct()

    # شرط وجود Serialnumber
    cars = cars.filter(serial__isnull=False).distinct()

    # شرط عدم وجود شکایت فعال (is_resolved=False)
    active_complaints = Complaint.objects.filter(is_resolved=False).values_list('car_id', flat=True)
    cars = cars.exclude(id__in=active_complaints).distinct()

    # شرط عدم وجود بازدید اول با exit=True یا accept=True
    completed_inspections = FirstInspection.objects.filter(
        Q(exit=True) | Q(accept=True)
    ).values_list('car_id', flat=True)
    cars = cars.exclude(id__in=completed_inspections).distinct()

    # هزینه‌های کامل با شرط agahi_type
    complete_costs = CarCosts.objects.filter(
        daily_price__isnull=False,
        signature_fee__isnull=False,
        notary_commitment__isnull=False,
        tax_fee__isnull=False,
        municipal_clearance_fee__isnull=False,
        highway_fee__isnull=False,
        transport_cost__isnull=False,
        signature_document__isnull=False,
        notary_document__isnull=False,
        tax_document__isnull=False,
        municipal_document__isnull=False,
        highway_document__isnull=False,
        transport_document__isnull=False,
        total_cost__isnull=False,
        recorded_by__isnull=False,
        recorded_at__isnull=False,
        company__isnull=False
    ).filter(
        Q(agahi_type__isnull=False, agahi_type__len__gt=0,
          agahi_cost__isnull=False,
          agahi_document__isnull=False) |
        Q(agahi_type__isnull=True) | Q(agahi_type__len=0)
    ).exclude(
        Q(agahi_type__isnull=False, agahi_type__len__gt=0) &
        (Q(agahi_cost__isnull=True) | Q(agahi_document=''))
    )

    # فیلتر خودروها بر اساس هزینه‌های کامل
    cars = cars.filter(costs__in=complete_costs).distinct()

    # جستجوی پیشرفته
    query = search_params.get('q', request.GET.get('q', ''))
    start_date = search_params.get('start_date', request.GET.get('start_date', ''))
    end_date = search_params.get('end_date', request.GET.get('end_date', ''))
    acceptance_number = search_params.get('acceptance_number', request.GET.get('acceptance_number', ''))
    has_inspection = search_params.get('has_inspection', request.GET.get('has_inspection', ''))
    engine_number = search_params.get('engine_number', request.GET.get('engine_number', ''))
    chassis_number = search_params.get('chassis_number', request.GET.get('chassis_number', ''))

    filtered_cars = list(cars)  # به لیست تبدیل می‌کنیم تا بتونیم فیلترهای پایتونی رو اعمال کنیم

    if query:
        filtered_cars = [car for car in filtered_cars if (
            query in car.acceptance_number or
            query in (car.license_plate or '') or
            query in (car.car_type or '')
        )]

    if start_date:
        try:
            start_date_gregorian = jdatetime_datetime.strptime(start_date, '%Y/%m/%d').togregorian().date()
            filtered_cars = [car for car in filtered_cars if car.delivery_date_gregorian >= start_date_gregorian]
        except ValueError:
            messages.error(request, "فرمت تاریخ شروع نامعتبر است.")

    if end_date:
        try:
            end_date_gregorian = jdatetime_datetime.strptime(end_date, '%Y/%m/%d').togregorian().date()
            filtered_cars = [car for car in filtered_cars if car.delivery_date_gregorian <= end_date_gregorian]
        except ValueError:
            messages.error(request, "فرمت تاریخ پایان نامعتبر است.")

    if acceptance_number:
        filtered_cars = [car for car in filtered_cars if acceptance_number in car.acceptance_number]

    if has_inspection in ['yes', 'no']:
        if has_inspection == 'yes':
            filtered_cars = [car for car in filtered_cars if car.first_inspections.exists()]
        elif has_inspection == 'no':
            filtered_cars = [car for car in filtered_cars if not car.first_inspections.exists()]

    if engine_number:
        filtered_cars = [car for car in filtered_cars if engine_number in (car.engine_number or '')]

    if chassis_number:
        filtered_cars = [car for car in filtered_cars if chassis_number in (car.chassis_number or '')]

    # بررسی درخواست برای خروجی اکسل
    if 'export_to_excel' in request.GET:
        print("درخواست اکسل دریافت شد")
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "لیست بازدید اول"
            headers = ['ردیف', 'سیستم', 'رنگ', 'شماره پذیرش', 'شماره انتظامی']
            for col_num, header in enumerate(headers, 1):
                ws[f'{get_column_letter(col_num)}1'] = header

            print(f"تعداد خودروها: {len(filtered_cars)}")
            for row_num, car in enumerate(filtered_cars, 2):
                ws[f'A{row_num}'] = row_num - 1
                ws[f'B{row_num}'] = car.car_system or ''
                ws[f'C{row_num}'] = car.car_color or ''
                ws[f'D{row_num}'] = car.acceptance_number or ''
                ws[f'E{row_num}'] = car.license_plate or ''

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            print(f"اندازه فایل: {len(output.getvalue())} بایت")

            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="first_inspection_list.xlsx"'
            return response
        except Exception as e:
            print(f"خطا: {str(e)}")
            return HttpResponse(f"خطا: {str(e)}", status=500)
    # صفحه‌بندی
    paginator = Paginator(filtered_cars, 10)
    page_number = request.GET.get('page')
    cars_page = paginator.get_page(page_number)

    # مدیریت درخواست POST برای دکمه "مدیریت"
    if request.method == 'POST':
        selected_car = request.POST.get('selected_car')
        if not selected_car:
            messages.error(request, "لطفاً یک خودرو را انتخاب کنید.")
        elif 'manage_first_inspection' in request.POST:
            request.session['return_to'] = request.get_full_path()
            return redirect('manage_first_inspection', acceptance_number=selected_car)

    return render(request, 'core/first_inspection_list.html', {
        'cars': cars_page,
        'request': request,
        'query': query,
        'start_date': start_date,
        'end_date': end_date,
        'acceptance_number': acceptance_number,
        'has_inspection': has_inspection,
        'engine_number': engine_number,
        'chassis_number': chassis_number,
    })


@login_required
def manage_first_inspection(request, acceptance_number, inspection_id=None):
    sys.stdout.reconfigure(encoding='utf-8')

    car = get_object_or_404(CarEntry, acceptance_number=acceptance_number)

    try:
        company = request.user.userprofile.company
        if not company:
            messages.error(request, "شما به شرکتی متصل نیستید.")
            return redirect('company_select')
        if not request.user.is_superuser and car.company != company:
            messages.error(request, "شما اجازه دسترسی به این خودرو را ندارید.")
            return redirect('first_inspection_list')
        first_inspections = FirstInspection.objects.filter(car=car, company=company).order_by(
            '-inspection_date_gregorian')
    except AttributeError:
        messages.error(request, "پروفایل شما کامل نیست.")
        return redirect('company_select')

    if inspection_id:
        first_inspection = get_object_or_404(FirstInspection, id=inspection_id, car=car, company=company)
    else:
        first_inspection = None

    # گرفتن آدرس کامل صفحه قبلی
    next_url = request.session.get('return_to', reverse('first_inspection_list'))

    if request.method == 'POST':
        if 'back' in request.POST:
            return redirect(next_url)

        form = FirstInspectionForm(request.POST, request.FILES, instance=first_inspection, request=request)
        if form.is_valid():
            first_inspection = form.save(commit=False)
            first_inspection.car = car
            first_inspection.company = company

            if 'save' in request.POST:
                first_inspection.accept = True
                first_inspection.save()
                messages.success(request,
                                 f"بازدید اول برای خودرو {car.acceptance_number} با موفقیت {'ویرایش و تأیید' if inspection_id else 'ثبت و تأیید'} شد.")
                return redirect(next_url)

            elif 'exit' in request.POST:
                first_inspection.exit = True
                car.is_active = True
                first_inspection.comment_exit = request.POST.get('comment_exit', '')
                if not first_inspection.comment_exit:
                    messages.error(request, "لطفاً علت خروج را وارد کنید.")
                    return render(request, 'core/manage_first_inspection.html', {
                        'form': form, 'car': car, 'first_inspections': first_inspections, 'inspection': first_inspection,
                        'next_url': next_url
                    })
                first_inspection.save()
                car.save()
                messages.success(request,
                                 f"خودرو {car.acceptance_number} از چرخه خارج {'و ویرایش' if inspection_id else ''} شد.")
                return redirect(next_url)

            elif 'stay_in_list' in request.POST:  # هماهنگ با نام دکمه توی HTML
                first_inspection.save()
                messages.success(request,
                                 f"بازدید اول برای خودرو {car.acceptance_number} با موفقیت {'ویرایش' if inspection_id else 'ثبت'} شد و در لیست باقی ماند.")
                return redirect(next_url)

        else:
            messages.error(request, "لطفاً خطاها را برطرف کنید.")
            print("Form errors:", form.errors)
    else:
        form = FirstInspectionForm(instance=first_inspection, request=request)

    return render(request, 'core/manage_first_inspection.html', {
        'form': form,
        'car': car,
        'first_inspections': first_inspections,
        'inspection': first_inspection,
        'next_url': next_url
    })
# @login_required
# def add_second_inspection(request, acceptance_number):
#     try:
#         car = CarEntry.objects.get(acceptance_number=acceptance_number)
#     except CarEntry.DoesNotExist:
#         messages.error(request, f"خودرو با شماره پذیرش {acceptance_number} یافت نشد.")
#         return redirect('home')
#
#     if not request.user.is_superuser and car.company != request.user.userprofile.company:
#         messages.error(request, "شما اجازه دسترسی به این خودرو را ندارید.")
#         return redirect('home')
#
#     if request.method == 'POST':
#         form = SecondInspectionForm(request.POST, request.FILES, request=request)
#         if form.is_valid():
#             second_inspection = form.save(commit=False)
#             second_inspection.car = car
#             second_inspection.recorded_by = request.user
#             second_inspection.company = car.company
#             second_inspection.save()
#             messages.success(request, f"بازدید دوم برای خودرو {car.acceptance_number} با موفقیت ثبت شد.")
#             return redirect('car_details', acceptance_number=car.acceptance_number)
#     else:
#         form = SecondInspectionForm(request=request)
#
#     return render(request, 'core/add_second_inspection.html', {
#         'form': form,
#         'car': car
#     })

@login_required
def second_inspection_list(request):
    sys.stdout.reconfigure(encoding='utf-8')

    # ذخیره پارامترهای جستجو توی سشن
    if request.method == 'GET' and any(key in request.GET for key in ['q', 'start_date', 'end_date', 'acceptance_number', 'has_inspection', 'engine_number', 'chassis_number']):
        request.session['search_params'] = {
            'q': request.GET.get('q', ''),
            'start_date': request.GET.get('start_date', ''),
            'end_date': request.GET.get('end_date', ''),
            'acceptance_number': request.GET.get('acceptance_number', ''),
            'has_inspection': request.GET.get('has_inspection', ''),
            'engine_number': request.GET.get('engine_number', ''),
            'chassis_number': request.GET.get('chassis_number', '')
        }
    elif 'reset_search' in request.GET:
        request.session['search_params'] = {}

    search_params = request.session.get('search_params', {})

    # فیلتر اولیه خودروها بر اساس شرکت کاربر
    if request.user.is_superuser:
        cars = CarEntry.objects.all()
    else:
        try:
            company = request.user.userprofile.company
            if not company:
                messages.error(request, "شما به شرکتی متصل نیستید.")
                return redirect('company_select')
            cars = CarEntry.objects.filter(company=company)
        except AttributeError:
            messages.error(request, "پروفایل شما کامل نیست.")
            return redirect('company_select')

    # اعمال شروط اولیه از first_inspection_list
    cars = cars.filter(
        documents__isnull=False,
        Physical_readiness=True
    ).distinct()

    # سندهای کامل
    complete_docs = Document.objects.filter(
        Q(is_deceased=True,
          car_document__isnull=False,
          vekalat__isnull=False,
          deceased_doc1__isnull=False,
          deceased_doc2__isnull=False) |
        Q(is_deceased=False,
          car_document__isnull=False,
          vekalat__isnull=False)
    ).exclude(
        Q(car_document='') |
        Q(vekalat='') |
        Q(is_deceased=True) & (Q(deceased_doc1='') | Q(deceased_doc2=''))
    )
    cars = cars.filter(document__in=complete_docs).distinct()

    # شرط وجود Serialnumber
    cars = cars.filter(serial__isnull=False).distinct()

    # شرط عدم وجود شکایت فعال
    active_complaints = Complaint.objects.filter(is_resolved=False).values_list('car_id', flat=True)
    cars = cars.exclude(id__in=active_complaints).distinct()

    # شرط بازدید اول با exit=False و accept=True
    cars = cars.filter(first_inspections__isnull=False).distinct()
    filtered_cars = []
    for car in cars:
        latest_first_inspection = car.first_inspections.order_by('-inspection_date_gregorian').first()
        if latest_first_inspection and not latest_first_inspection.exit and latest_first_inspection.accept:
            filtered_cars.append({
                'car': car,
                'latest_inspection': latest_first_inspection
            })

    # هزینه‌های کامل با شرط agahi_type
    complete_costs = CarCosts.objects.filter(
        daily_price__isnull=False,
        signature_fee__isnull=False,
        notary_commitment__isnull=False,
        tax_fee__isnull=False,
        municipal_clearance_fee__isnull=False,
        highway_fee__isnull=False,
        transport_cost__isnull=False,
        signature_document__isnull=False,
        notary_document__isnull=False,
        tax_document__isnull=False,
        municipal_document__isnull=False,
        highway_document__isnull=False,
        transport_document__isnull=False,
        total_cost__isnull=False,
        recorded_by__isnull=False,
        recorded_at__isnull=False,
        company__isnull=False
    ).filter(
        Q(agahi_type__isnull=False, agahi_type__len__gt=0,
          agahi_cost__isnull=False,
          agahi_document__isnull=False) |
        Q(agahi_type__isnull=True) | Q(agahi_type__len=0)
    ).exclude(
        Q(agahi_type__isnull=False, agahi_type__len__gt=0) &
        (Q(agahi_cost__isnull=True) | Q(agahi_document=''))
    )
    cars = cars.filter(costs__in=complete_costs).distinct()

    # جستجوی پیشرفته
    query = search_params.get('q', request.GET.get('q', ''))
    start_date = search_params.get('start_date', request.GET.get('start_date', ''))
    end_date = search_params.get('end_date', request.GET.get('end_date', ''))
    acceptance_number = search_params.get('acceptance_number', request.GET.get('acceptance_number', ''))
    has_inspection = search_params.get('has_inspection', request.GET.get('has_inspection', ''))
    engine_number = search_params.get('engine_number', request.GET.get('engine_number', ''))
    chassis_number = search_params.get('chassis_number', request.GET.get('chassis_number', ''))

    if query:
        filtered_cars = [item for item in filtered_cars if (
            query in item['car'].acceptance_number or
            query in (item['car'].license_plate or '') or
            query in (item['car'].car_type or '')
        )]

    if start_date:
        try:
            start_date_gregorian = jdatetime_datetime.strptime(start_date, '%Y/%m/%d').togregorian().date()
            filtered_cars = [item for item in filtered_cars if item['car'].delivery_date_gregorian >= start_date_gregorian]
        except ValueError:
            messages.error(request, "فرمت تاریخ شروع نامعتبر است.")

    if end_date:
        try:
            end_date_gregorian = jdatetime_datetime.strptime(end_date, '%Y/%m/%d').togregorian().date()
            filtered_cars = [item for item in filtered_cars if item['car'].delivery_date_gregorian <= end_date_gregorian]
        except ValueError:
            messages.error(request, "فرمت تاریخ پایان نامعتبر است.")

    if acceptance_number:
        filtered_cars = [item for item in filtered_cars if acceptance_number in item['car'].acceptance_number]

    if has_inspection in ['yes', 'no']:
        if has_inspection == 'yes':
            filtered_cars = [item for item in filtered_cars if item['car'].first_inspections.exists()]
        elif has_inspection == 'no':
            filtered_cars = [item for item in filtered_cars if not item['car'].first_inspections.exists()]

    if engine_number:
        filtered_cars = [item for item in filtered_cars if engine_number in (item['car'].engine_number or '')]

    if chassis_number:
        filtered_cars = [item for item in filtered_cars if chassis_number in (item['car'].chassis_number or '')]

        # خروجی اکسل
    if 'export_to_excel' in request.GET:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "لیست بازدید دوم"

        # ستون‌ها
        headers = ['شماره سریال', 'شماره انتظامی', 'تاریخ بازدید اول', 'شرکت', 'شناسه ملی']
        worksheet.append(headers)

        # پر کردن داده‌ها
        for item in filtered_cars:
            car = item['car']
            latest_inspection = item['latest_inspection']
            serial_number = car.serial.serial_number if hasattr(car, 'serial') else '-'
            license_plate = car.license_plate or '-'
            inspection_date = latest_inspection.inspection_date or '-'
            company_name = car.company.name if car.company else '-'
            national_id = car.company.National_ID if car.company else '-'

            worksheet.append([
                serial_number,
                license_plate,
                inspection_date,
                company_name,
                national_id
            ])

        # تنظیم پاسخ HTTP برای دانلود فایل
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="second_inspection_list.xlsx"'
        workbook.save(response)
        return response

    # صفحه‌بندی
    paginator = Paginator(filtered_cars, 10)
    page_number = request.GET.get('page')
    cars_page = paginator.get_page(page_number)

    # مدیریت عملیات POST
    if request.method == 'POST':
        selected_car = request.POST.get('selected_car')
        if selected_car:
            if 'return_to' in request.session:
                del request.session['return_to']  # پاک کردن مقدار قدیمی
            request.session['return_to'] = request.get_full_path()
            print("Stored return_to in second_inspection_list:", request.session['return_to'])
            if 'finalize_registration' in request.POST:
                return redirect('finalize_registration', acceptance_number=selected_car)
        else:
            messages.error(request, "لطفاً یک خودرو انتخاب کنید.")

    return render(request, 'core/second_inspection_list.html', {
        'cars': cars_page,
        'request': request,
        'query': query,
        'start_date': start_date,
        'end_date': end_date,
        'acceptance_number': acceptance_number,
        'has_inspection': has_inspection,
        'engine_number': engine_number,
        'chassis_number': chassis_number,
    })

@login_required
def finalized_list(request):
    cars = CarEntry.objects.filter(final_registration__isnull=False).order_by('-final_registration__finalized_at')
    # کپی کد registration_list برای فیلتر و صفحه‌بندی
    query = request.GET.get('q', '').strip()
    if query:
        cars = cars.filter(
            Q(acceptance_number__icontains=query) |
            Q(license_plate__icontains=query) |
            Q(car_type__icontains=query) |
            Q(owner_name__icontains=query)
        )

    paginator = Paginator(cars, 10)
    page_number = request.GET.get('page')
    try:
        page_cars = paginator.page(page_number)
    except PageNotAnInteger:
        page_cars = paginator.page(1)
    except EmptyPage:
        page_cars = paginator.page(paginator.num_pages)

    return render(request, 'core/finalized_list.html', {
        'cars': page_cars,
    })


@login_required
def connect_car_document(request):
    if request.method == 'POST':
        engine_number = request.POST.get('engine_number')
        chassis_number = request.POST.get('chassis_number')

        if not engine_number or not chassis_number:
            messages.error(request, 'شماره موتور و شماره شاسی نمی‌توانند خالی باشند.')
            return render(request, 'core/connect_car_document.html', {})

        try:
            car = CarEntry.objects.get(engine_number=engine_number, chassis_number=chassis_number)
            doc = Document.objects.get(engine_number=engine_number, chassis_number=chassis_number)

            if car.document or doc.car:
                messages.error(request, 'این خودرو یا سند قبلاً به هم متصل شده‌اند.')
            else:
                with transaction.atomic():
                    car.document = doc
                    doc.car = car
                    car.save(update_fields=['document'])
                    doc.save(update_fields=['car'])
                messages.success(request, f'خودرو {car.acceptance_number} با سند {doc.id} با موفقیت متصل شد.')
                return redirect('home')  # یا هر URL دلخواه

        except CarEntry.DoesNotExist:
            messages.error(request, f'خودرویی با شماره موتور {engine_number} و شاسی {chassis_number} پیدا نشد.')
        except Document.DoesNotExist:
            messages.error(request, f'سندی با شماره موتور {engine_number} و شاسی {chassis_number} پیدا نشد.')
        except CarEntry.MultipleObjectsReturned:
            messages.error(request, f'چند خودرو با شماره موتور {engine_number} و شاسی {chassis_number} پیدا شد.')
        except Document.MultipleObjectsReturned:
            messages.error(request, f'چند سند با شماره موتور {engine_number} و شاسی {chassis_number} پیدا شد.')
        except Exception as e:
            messages.error(request, f'خطایی رخ داد: {str(e)}')

    return render(request, 'core/connect_car_document.html', {})

@login_required
def view_signal_logs(request):
    log_file_path = os.path.join(os.path.dirname(__file__), '..', 'signals.log')  # مسیر فایل لاگ
    logs = []
    if os.path.exists(log_file_path):
        with open(log_file_path, 'r', encoding='utf-8') as log_file:
            logs = log_file.readlines()
    return render(request, 'core/view_signal_logs.html', {'logs': logs})